"""
CAD Digit OCR – Flask Backend
==============================
Serves the REST API consumed by the React frontend.

Changes from original:
  - CNN predict_image() replaced with EasyOCR (fixes overconfidence + domain mismatch)
  - EasyOCR runs full 8-rotation sweep per crop (0,90,180,270,45,135,225,315)
    with fast-accept at >= 0.85 confidence on cardinal angles
  - run_pipeline() now reports sub-step progress (extract/cluster/candidates/ocr)
  - Live digit counter + ETA written to state after every single crop
  - api_check_model() updated — no longer requires cad_digit_model.pt
  - _prewarm_trocr() replaced with _prewarm_ocr() for EasyOCR
  - Dead CNN code (load_model, predict_image, val_transform) removed
  - UPDATED: Arrays support for multi-layer OCR processing and multi-layer Cable span building
  - UPDATED: Hatch Support added for frontend rendering.
"""

import argparse
import atexit
import base64
import io
import json
import math
import os
import socket
import shutil as _shutil
import subprocess
import sys
import tempfile
import threading
import time
import uuid
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import cv2
import ezdxf
import numpy as np
from flask import Blueprint, Flask, jsonify, request, send_file, send_from_directory
from flask_cors import CORS
from PIL import Image

from app_python.planner_config import DEFAULT_PROJECT_ID, ENABLE_PLANNER_INTEGRATION
from app_python.services import span_builder
from app_python.services.planner_auth import auth
from app_python.services.session_store import (
    compute_checksum,
    get_or_create_project_session,
    save_full_results,
    session_has_user_edits,
)

app = Flask(__name__, static_folder="frontend/dist", static_url_path="")
CORS(app)

public_api = Blueprint("public_api", __name__, url_prefix="/api/v1")


# ─────────────────────────────────────────────────────────────────────────────
# PDF → DXF  (AutoCAD)
# ─────────────────────────────────────────────────────────────────────────────


def pdf_to_dxf_autocad(pdf_path):
    pdf_path = Path(pdf_path).resolve()
    dxf_path = pdf_path.with_suffix(".dxf")

    accore_candidates = [
        r"C:\Program Files\Autodesk\AutoCAD 2027\accoreconsole.exe",
        r"C:\Program Files\Autodesk\AutoCAD 2026\accoreconsole.exe",
        r"C:\Program Files\Autodesk\AutoCAD 2025\accoreconsole.exe",
        r"C:\Program Files\Autodesk\AutoCAD 2024\accoreconsole.exe",
        r"C:\Program Files\Autodesk\AutoCAD 2023\accoreconsole.exe",
        r"C:\Program Files\Autodesk\AutoCAD 2022\accoreconsole.exe",
        r"C:\Program Files\Autodesk\AutoCAD 2021\accoreconsole.exe",
    ]
    accore = next((p for p in accore_candidates if Path(p).exists()), None)
    if accore is None:
        raise RuntimeError(
            "AutoCAD is not installed or not found. "
            "PDF conversion requires AutoCAD 2022 or later. "
            "Please convert your PDF to DXF manually and upload the DXF file instead."
        )

    script = f'''FILEDIA 0
-PDFIMPORT
FILE
"{pdf_path}"
1
0,0
1
0
DXFOUT
"{dxf_path}"
16
Version
2018
QUIT
'''
    with tempfile.NamedTemporaryFile(delete=False, suffix=".scr", mode="w") as f:
        f.write(script)
        script_path = f.name

    result = subprocess.run(
        [str(accore), "/s", script_path],
        check=True,
        capture_output=True,
        text=True,
        timeout=120,
    )

    if not dxf_path.exists():
        raise RuntimeError("DXF was not created by AutoCAD.")

    return str(dxf_path)


# ─────────────────────────────────────────────────────────────────────────────
# DXF PIPELINE CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────

CONNECT_TOL = 0.20
MIN_TOTAL_LENGTH = 0.15
EPS_THIN = 0.03
LONG_DIM = 3.0
COMPLEX_MIN = 0.15
MIN_SEGS_FOR_DIGIT = 2
MAX_DOM_DIR = 0.88
MAX_ENDPOINTS_FOR_LINE = 2
ENDPOINT_TOL_SCALE = 1.0
W_FACTOR = 6.0
H_FACTOR = 6.0
LEN_FACTOR = 10.0
AREA_FACTOR = 18.0
MAX_ASPECT = 8.0
SALVAGE_DIST_FACTOR = 0.25
SALVAGE_LONG_SEG_FACTOR = 1.8
SALVAGE_ANGLE_TOL_DEG = 20.0
CABLE_CONNECT_TOL = 0.10

# ── adaptive scale ────────────────────────────────────────────────────────────
# The absolute thresholds above (CONNECT_TOL, MIN_TOTAL_LENGTH, EPS_THIN,
# LONG_DIM) were calibrated for drawings whose median stroke length is
# REF_MEDIAN_SEGLEN. Different maps are drawn at different scales, so instead of
# hand-editing CONNECT_TOL per map we measure each drawing's own stroke size and
# scale every distance threshold by the same factor. estimate_scale() returns
# that multiplier (1.0 == reference scale). Overridable via env for tuning.
REF_MEDIAN_SEGLEN = float(os.environ.get("STRAND_REF_SEGLEN", "0.0125"))


def estimate_scale(segments) -> float:
    """Per-drawing scale multiplier = median stroke length / reference length.

    Robust to map scale: a drawing drawn 10x larger has ~10x longer strokes and
    therefore a ~10x larger connect tolerance, digit-length floor, etc. Cable
    polylines are a minority of segments, so the median tracks the digit strokes.
    """
    lens = [s.length() for s in segments if s.length() > 1e-9]
    if not lens:
        return 1.0
    return float(np.median(lens)) / REF_MEDIAN_SEGLEN


# A segment longer than this multiple of the median stroke length is a cable
# "seed" — clearly part of the strand line, not a digit. Digit strokes top out
# around ~10x median while cable seeds are ~15-120x median (empty valley between).
CABLE_SEG_FACTOR = 20.0


def cable_segment_indices(segments, scale=None) -> set:
    """Indices of segments forming the cable strand line, excluded from digit
    clustering so the cable can't swallow digits or leak a stub into a crop.

    The cable strand is drawn in a distinct ACI color (e.g. 254) covering both
    its long runs AND its short corner pieces, while digits are a different color
    (usually BYLAYER). We auto-detect the cable color as the dominant color among
    the over-long "seed" segments, then exclude EVERY segment of that color. This
    cleanly drops cable corner fragments (no junk candidates) while keeping digits
    that sit on the cable (they are a different color), which geometry alone
    cannot do.

    Falls back to removing just the long seeds when color is not discriminative
    (cable shares the digits' color), so it never removes a whole drawing.
    """
    from collections import Counter

    # Toggle: set STRAND_CABLE_SEPARATION=0 to disable cable exclusion entirely
    # (digits on the cable stay missing, but no cable-fragment false positives).
    if os.environ.get("STRAND_CABLE_SEPARATION", "1") == "0":
        return set()

    lens = [s.length() for s in segments if s.length() > 1e-9]
    if not lens:
        return set()
    med = float(np.median(lens))
    seeds = [i for i, s in enumerate(segments) if s.length() > CABLE_SEG_FACTOR * med]
    if not seeds:
        return set()

    cable_color, _ = Counter(segments[i].color for i in seeds).most_common(1)[0]
    same_color = {i for i, s in enumerate(segments) if s.color == cable_color}
    # Use color only when the cable color is a clear minority (otherwise it is
    # also the digit color and excluding it would wipe out real digits).
    if len(same_color) <= 0.6 * len(segments):
        return same_color
    return set(seeds)


# ─────────────────────────────────────────────────────────────────────────────
# DATA CLASSES
# ─────────────────────────────────────────────────────────────────────────────


@dataclass
class Seg:
    x1: float
    y1: float
    x2: float
    y2: float
    is_hatch: bool = False  # Added flag to identify hatch boundaries
    color: int = 256  # ACI color (256 = BYLAYER); used to tell cable from digits

    def p1(self):
        return (self.x1, self.y1)

    def p2(self):
        return (self.x2, self.y2)

    def length(self):
        return math.hypot(self.x2 - self.x1, self.y2 - self.y1)


@dataclass
class ClusterInfo:
    cluster_id: int
    seg_indices: List[int]
    bbox: Tuple[float, float, float, float]
    width: float
    height: float
    total_length: float
    kind: str


@dataclass
class Candidate:
    digit_id: int
    cluster_id: int
    seg_indices: List[int]
    bbox: Tuple[float, float, float, float]
    width: float
    height: float
    total_length: float


# ─────────────────────────────────────────────────────────────────────────────
# EASYOCR SINGLETON + MULTI-ROTATION INFERENCE
# ─────────────────────────────────────────────────────────────────────────────

import re as _re

_easyocr_reader = None
_easyocr_lock = threading.Lock()

_OCR_ROTATION_PAIRS = [
    (0, 180),
    (90, 270),
    (15, 195),
    (30, 210),
    (45, 225),
    (60, 240),
    (75, 255),
    (105, 285),
    (120, 300),
    (135, 315),
    (150, 330),
    (165, 345),
]

_FAST_ACCEPT_CONF = 0.95
# Confidence from strand_recognizer is a vote-share / agreement score (0-1),
# not a raw single-pass softmax. 0.50 means "the winning value carried at least
# a majority of the weighted votes"; below that we flag for human review.
_MIN_CONF = 0.50
# Strand lengths reach ~500 m on backbone drawings; mirror strand_recognizer so
# the two gates cannot disagree about what a valid value is.
from app_python.services.strand_recognizer import STRAND_MAX as _MAX_STRAND_VALUE
_STRAND_RE = _re.compile(r"^\d{1,3}$")


def _is_valid_strand(text: str) -> bool:
    if not text or not _STRAND_RE.match(text):
        return False
    try:
        return int(text) <= _MAX_STRAND_VALUE
    except ValueError:
        return False


def _load_easyocr():
    global _easyocr_reader
    if _easyocr_reader is not None:
        return _easyocr_reader

    with _easyocr_lock:
        if _easyocr_reader is not None:
            return _easyocr_reader

        import easyocr
        import torch

        print("[ocr] Loading EasyOCR reader (first-time download may take ~30s)...")
        use_gpu = bool(torch.cuda.is_available())
        print(f"[ocr] EasyOCR device: {'cuda' if use_gpu else 'cpu'}")
        _easyocr_reader = easyocr.Reader(["en"], gpu=use_gpu)
        print("[ocr] EasyOCR ready.")

    return _easyocr_reader


def _rotate_crop(img: np.ndarray, degrees: int) -> np.ndarray:
    if degrees == 0:
        return img
    if degrees == 90:
        return cv2.rotate(img, cv2.ROTATE_90_CLOCKWISE)
    if degrees == 180:
        return cv2.rotate(img, cv2.ROTATE_180)
    if degrees == 270:
        return cv2.rotate(img, cv2.ROTATE_90_COUNTERCLOCKWISE)

    from PIL import Image as _PILImage

    # expand=True grows the canvas to fit the rotated content; pad to square
    # so the caller always gets a square image (no squish back to orig dims).
    pil = _PILImage.fromarray(img)
    rotated = pil.rotate(-degrees, resample=_PILImage.BICUBIC, expand=True, fillcolor=0)
    rw, rh = rotated.size
    side = max(rw, rh)
    sq = _PILImage.new("L", (side, side), 0)
    sq.paste(rotated, ((side - rw) // 2, (side - rh) // 2))
    return np.array(sq)


def _easyocr_on_prepared(img: np.ndarray) -> Tuple[str, float]:
    reader = _load_easyocr()
    h, w = img.shape[:2]

    try:
        results = reader.recognize(
            img,
            horizontal_list=[[0, w, 0, h]],
            free_list=[],
            allowlist="0123456789",
            detail=1,
        )
    except Exception:
        results = reader.readtext(
            img,
            allowlist="0123456789",
            detail=1,
            paragraph=False,
        )

    if not results:
        return "", 0.0

    best = max(results, key=lambda x: x[2])
    text = best[1].strip()
    conf = float(best[2])

    if not _is_valid_strand(text):
        return text, conf * 0.5

    return text, conf


def predict_with_easyocr(crop_np: np.ndarray) -> Tuple[str, float]:
    """
    Delegate to strand_recognizer: deskew, multi-variant rendering, and
    domain-constrained voting, EasyOCR only. TrOCR is opt-in via
    STRAND_USE_TROCR=1 and off by default — there is no dual-engine ensemble.
    """
    from app_python.services.strand_recognizer import recognize
    return recognize(crop_np)


# ─────────────────────────────────────────────────────────────────────────────
# DXF GEOMETRY HELPERS
# ─────────────────────────────────────────────────────────────────────────────


def _dist2(a, b):
    return (a[0] - b[0]) ** 2 + (a[1] - b[1]) ** 2


def point_segment_dist(px, py, x1, y1, x2, y2):
    """Calculates the shortest distance from a point to a line segment."""
    l2 = (x2 - x1) ** 2 + (y2 - y1) ** 2
    if l2 == 0:
        return math.hypot(px - x1, py - y1)
    t = max(0.0, min(1.0, ((px - x1) * (x2 - x1) + (py - y1) * (y2 - y1)) / l2))
    proj_x = x1 + t * (x2 - x1)
    proj_y = y1 + t * (y2 - y1)
    return math.hypot(px - proj_x, py - proj_y)


def _bbox_from_segments(segments, idxs):
    xs, ys = [], []
    for i in idxs:
        s = segments[i]
        xs += [s.x1, s.x2]
        ys += [s.y1, s.y2]
    return (min(xs), min(ys), max(xs), max(ys))


def _segmentize(pts, closed):
    segs = []
    for i in range(len(pts) - 1):
        segs.append(Seg(pts[i][0], pts[i][1], pts[i + 1][0], pts[i + 1][1]))
    if closed and len(pts) > 2:
        segs.append(Seg(pts[-1][0], pts[-1][1], pts[0][0], pts[0][1]))
    return segs


def list_layers(dxf_path):
    doc = ezdxf.readfile(dxf_path)
    return sorted(layer.dxf.name for layer in doc.layers)


def extract_stroke_segments(doc, layer_name, include_circles=True):
    segments = []
    ARC_STEPS, SPLINE_STEPS, CIRCLE_STEPS = 24, 30, 36

    def iter_spaces():
        yield doc.modelspace()
        for layout in doc.layouts:
            if layout.name.lower() != "model":
                yield layout

    for space in iter_spaces():
        for e in space:
            if getattr(e.dxf, "layer", None) != layer_name:
                continue
            # Tag every segment from this entity with its ACI color so the cable
            # strand line (a distinct color) can be told apart from the digits.
            try:
                col = int(e.dxf.color)
            except Exception:
                col = 256
            seg_start = len(segments)
            t = e.dxftype()
            if t == "LINE":
                p1, p2 = e.dxf.start, e.dxf.end
                segments.append(Seg(float(p1.x), float(p1.y), float(p2.x), float(p2.y)))
            elif t == "LWPOLYLINE":
                pts = [(float(x), float(y)) for x, y, *_ in e.get_points("xy")]
                segments.extend(_segmentize(pts, bool(e.closed)))
            elif t == "POLYLINE":
                pts = [
                    (float(v.dxf.location.x), float(v.dxf.location.y))
                    for v in e.vertices
                ]
                segments.extend(_segmentize(pts, bool(e.is_closed)))
            elif t == "ARC":
                c, r = e.dxf.center, float(e.dxf.radius)
                a0 = math.radians(float(e.dxf.start_angle))
                a1 = math.radians(float(e.dxf.end_angle))
                if a1 < a0:
                    a1 += 2 * math.pi
                angles = np.linspace(a0, a1, ARC_STEPS)
                pts = [
                    (float(c.x) + r * math.cos(a), float(c.y) + r * math.sin(a))
                    for a in angles
                ]
                segments.extend(_segmentize(pts, False))
            elif t == "CIRCLE":
                # This parameter was declared but never read — six call sites
                # passed include_circles=False in vain (the old check compared
                # against POLE_LAYER_FILTER, whose entry "pole, stp" matches no
                # real layer name).
                if not include_circles:
                    continue
                c, r = e.dxf.center, float(e.dxf.radius)
                angles = np.linspace(0, 2 * math.pi, CIRCLE_STEPS, endpoint=False)
                pts = [
                    (float(c.x) + r * math.cos(a), float(c.y) + r * math.sin(a))
                    for a in angles
                ]
                segments.extend(_segmentize(pts, True))
            elif t == "SPLINE":
                try:
                    from ezdxf.math import BSpline

                    bs = BSpline.from_spline(e)
                    pts = [
                        (float(p.x), float(p.y))
                        for p in (bs.point(t) for t in np.linspace(0, 1, SPLINE_STEPS))
                    ]
                    segments.extend(_segmentize(pts, False))
                except Exception:
                    pass
            elif t == "HATCH":
                # Extract hatch boundaries as segments so they render on the frontend
                try:
                    from ezdxf.path import from_hatch

                    for p in from_hatch(e):
                        pts = list(p.flattening(distance=0.1))
                        if len(pts) > 1:
                            for i in range(len(pts) - 1):
                                segments.append(
                                    Seg(
                                        float(pts[i].x),
                                        float(pts[i].y),
                                        float(pts[i + 1].x),
                                        float(pts[i + 1].y),
                                        is_hatch=True,
                                    )
                                )
                            if (
                                abs(pts[0].x - pts[-1].x) > 1e-6
                                or abs(pts[0].y - pts[-1].y) > 1e-6
                            ):
                                segments.append(
                                    Seg(
                                        float(pts[-1].x),
                                        float(pts[-1].y),
                                        float(pts[0].x),
                                        float(pts[0].y),
                                        is_hatch=True,
                                    )
                                )
                except Exception:
                    pass
            for s in segments[seg_start:]:
                s.color = col
    return segments


def cluster_segments(segments, tol=None, ignore=None):
    if not segments:
        return []
    if tol is None:
        tol = CONNECT_TOL * estimate_scale(segments)
    ignore = ignore or set()
    tol2 = tol * tol
    cell_size = tol

    def cell_key(p):
        return (int(math.floor(p[0] / cell_size)), int(math.floor(p[1] / cell_size)))

    grid = {}
    endpoints = [
        (i, s.p1()) for i, s in enumerate(segments) if i not in ignore
    ] + [(i, s.p2()) for i, s in enumerate(segments) if i not in ignore]
    for si, p in endpoints:
        grid.setdefault(cell_key(p), []).append((si, p))
    adj = [[] for _ in range(len(segments))]
    for si, p in endpoints:
        ck = cell_key(p)
        for dx in (-1, 0, 1):
            for dy in (-1, 0, 1):
                for sj, q in grid.get((ck[0] + dx, ck[1] + dy), []):
                    if sj != si and _dist2(p, q) <= tol2:
                        adj[si].append(sj)
    # Ignored (cable) segments never seed or join a cluster.
    visited = [(i in ignore) for i in range(len(segments))]
    clusters = []
    for i in range(len(segments)):
        if visited[i]:
            continue
        stack = [i]
        visited[i] = True
        comp = []
        while stack:
            cur = stack.pop()
            comp.append(cur)
            for nb in adj[cur]:
                if not visited[nb]:
                    visited[nb] = True
                    stack.append(nb)
        clusters.append(comp)
    return clusters


def cluster_complexity(segments, idxs):
    ang = []
    for i in idxs:
        s = segments[i]
        dx = s.x2 - s.x1
        dy = s.y2 - s.y1
        if abs(dx) < 1e-12 and abs(dy) < 1e-12:
            continue
        a = math.atan2(dy, dx)
        a = (a + math.pi) % math.pi
        ang.append(a)
    if len(ang) <= 1:
        return 0.0
    c = sum(math.cos(2 * a) for a in ang) / len(ang)
    s = sum(math.sin(2 * a) for a in ang) / len(ang)
    return 1.0 - math.hypot(c, s)


def endpoint_count(segments, idxs, tol):
    cell = tol

    def key(p):
        return (int(math.floor(p[0] / cell)), int(math.floor(p[1] / cell)))

    counts = {}
    for i in idxs:
        s = segments[i]
        for p in (s.p1(), s.p2()):
            k = key(p)
            counts[k] = counts.get(k, 0) + 1
    return sum(1 for v in counts.values() if v == 1)


def dominant_direction_ratio(segments, idxs, bins=12):
    hist = [0] * bins
    n = 0
    for i in idxs:
        s = segments[i]
        dx = s.x2 - s.x1
        dy = s.y2 - s.y1
        if abs(dx) < 1e-12 and abs(dy) < 1e-12:
            continue
        a = math.atan2(dy, dx)
        a = (a + math.pi) % math.pi
        b = int((a / math.pi) * bins) % bins
        hist[b] += 1
        n += 1
    return max(hist) / n if n > 0 else 1.0


def is_renderable_cluster(segments, info):
    if info.width * info.height < 1e-8:
        return False
    return any(segments[si].length() > 1e-6 for si in info.seg_indices)


def analyze_clusters(segments, clusters, scale=None):
    infos = []
    sc = scale if scale is not None else estimate_scale(segments)
    ep_tol = CONNECT_TOL * sc * ENDPOINT_TOL_SCALE
    min_total = MIN_TOTAL_LENGTH * sc
    eps_thin = EPS_THIN * sc
    long_dim = LONG_DIM * sc
    for cid, idxs in enumerate(clusters):
        minx, miny, maxx, maxy = _bbox_from_segments(segments, idxs)
        w = maxx - minx
        h = maxy - miny
        total_len = sum(segments[i].length() for i in idxs)
        if total_len < min_total:
            continue
        comp = cluster_complexity(segments, idxs)
        dom = dominant_direction_ratio(segments, idxs)
        ep = endpoint_count(segments, idxs, tol=ep_tol)
        thin = min(w, h) < eps_thin
        longish = max(w, h) > long_dim
        few = len(idxs) < MIN_SEGS_FOR_DIGIT
        if (
            (thin and longish and comp < COMPLEX_MIN)
            or (dom > MAX_DOM_DIR and ep <= MAX_ENDPOINTS_FOR_LINE)
            or (few and comp < COMPLEX_MIN)
        ):
            kind = "line"
        else:
            kind = "digit_candidate"
        infos.append(
            ClusterInfo(cid, idxs, (minx, miny, maxx, maxy), w, h, total_len, kind)
        )
    return infos


def _pca_main_axis(points):
    c = points.mean(axis=0)
    X = points - c
    C = (X.T @ X) / max(1, len(points) - 1)
    vals, vecs = np.linalg.eigh(C)
    d = vecs[:, np.argmax(vals)]
    return c, d / max(np.linalg.norm(d), 1e-12)


def _point_line_dist(p, c, d):
    v = p - c
    proj = np.dot(v, d) * d
    return float(np.linalg.norm(v - proj))


def salvage_remove_dominant_line(
    segments, idxs, connect_tol, dist_factor, long_seg_factor, angle_tol_deg,
    min_total=MIN_TOTAL_LENGTH,
):
    if len(idxs) < 4:
        return [idxs]
    pts = []
    seg_lens = []
    for i in idxs:
        s = segments[i]
        pts += [[s.x1, s.y1], [s.x2, s.y2]]
        seg_lens.append(s.length())
    pts = np.array(pts, dtype=float)
    seg_lens = np.array(seg_lens, dtype=float)
    minx, miny, maxx, maxy = _bbox_from_segments(segments, idxs)
    thin_dim = max(1e-9, min(maxx - minx, maxy - miny))
    c, d = _pca_main_axis(pts)
    med_len = float(np.median(seg_lens))
    long_thr = max(med_len * long_seg_factor, med_len + 1e-9)
    ang_tol = math.radians(angle_tol_deg)
    keep = []
    removed = []
    for i in idxs:
        s = segments[i]
        L = s.length()
        mid = np.array([(s.x1 + s.x2) * 0.5, (s.y1 + s.y2) * 0.5], dtype=float)
        dist = _point_line_dist(mid, c, d)
        v = np.array([s.x2 - s.x1, s.y2 - s.y1], dtype=float)
        nv = np.linalg.norm(v)
        if nv < 1e-12:
            keep.append(i)
            continue
        cosang = float(abs(np.dot(v / nv, d)))
        cosang = max(-1.0, min(1.0, cosang))
        ang = math.acos(cosang)
        if dist <= dist_factor * thin_dim and ang <= ang_tol and L >= long_thr:
            removed.append(i)
        else:
            keep.append(i)
    if not removed:
        return [idxs]
    kept_segs = [segments[i] for i in keep]
    subclusters_local = cluster_segments(kept_segs, tol=connect_tol)
    subclusters = [[keep[j] for j in comp] for comp in subclusters_local]
    subclusters = [
        c
        for c in subclusters
        if sum(segments[i].length() for i in c) >= min_total
    ]
    return subclusters if subclusters else [idxs]


def _split_by_gap(segments, info, med_w, med_h, min_total=MIN_TOTAL_LENGTH):
    mids_x = []
    mids_y = []
    for i in info.seg_indices:
        s = segments[i]
        mids_x.append((s.x1 + s.x2) / 2.0)
        mids_y.append((s.y1 + s.y2) / 2.0)

    if len(mids_x) < 4:
        return [info]

    def find_best_gap(values, total_dim):
        sv = sorted(values)
        best_gap = -1.0
        best_split = None
        for k in range(1, len(sv)):
            gap = sv[k] - sv[k - 1]
            if gap > best_gap:
                best_gap = gap
                best_split = (sv[k] + sv[k - 1]) / 2.0
        if best_split is None or best_gap < total_dim * 0.20:
            return None, -1.0
        return best_split, best_gap

    def do_split(axis_values, axis, threshold):
        a_idxs, b_idxs = [], []
        for k, i in enumerate(info.seg_indices):
            if axis_values[k] < threshold:
                a_idxs.append(i)
            else:
                b_idxs.append(i)
        results = []
        for half in (a_idxs, b_idxs):
            if len(half) < 2:
                continue
            bx = _bbox_from_segments(segments, half)
            w = bx[2] - bx[0]
            h = bx[3] - bx[1]
            tlen = sum(segments[j].length() for j in half)
            if tlen < min_total or w < 1e-9 or h < 1e-9:
                continue
            results.append(
                ClusterInfo(info.cluster_id, half, bx, w, h, tlen, "digit_candidate")
            )
        return results if len(results) == 2 else [info]

    if info.height > med_h * 1.5:
        split_y, gap_y = find_best_gap(mids_y, info.height)
        if split_y is not None:
            result = do_split(mids_y, "y", split_y)
            if len(result) == 2:
                return result

    if info.width > med_w * 3.0:
        split_x, gap_x = find_best_gap(mids_x, info.width)
        if split_x is not None:
            result = do_split(mids_x, "x", split_x)
            if len(result) == 2:
                return result

    return [info]


def build_candidates_robust(segments, infos, scale=None):
    sc = scale if scale is not None else estimate_scale(segments)
    min_total = MIN_TOTAL_LENGTH * sc
    prelim = [
        i
        for i in infos
        if i.kind == "digit_candidate" and is_renderable_cluster(segments, i)
    ]
    if not prelim:
        return []
    areas = np.array([i.width * i.height for i in prelim], dtype=float)
    cutoff = np.quantile(areas, 0.80)
    small = [i for i in prelim if i.width * i.height <= cutoff]
    base = small if len(small) >= 10 else prelim
    med_w = float(np.median([i.width for i in base]))
    med_h = float(np.median([i.height for i in base]))
    med_len = float(np.median([i.total_length for i in base]))
    med_area = float(np.median([i.width * i.height for i in base]))

    def aspect_ok(w, h):
        return (
            max(w, h) / max(min(w, h), 1e-12) <= MAX_ASPECT
            if w > 1e-12 and h > 1e-12
            else False
        )

    final_infos = []
    for i in prelim:
        area = i.width * i.height
        too_big = (
            i.width > med_w * W_FACTOR
            or i.height > med_h * H_FACTOR
            or i.total_length > med_len * LEN_FACTOR
            or area > med_area * AREA_FACTOR
            or not aspect_ok(i.width, i.height)
        )
        if not too_big:
            for split in _split_by_gap(segments, i, med_w, med_h, min_total=min_total):
                final_infos.append(split)
            continue
        subclusters = salvage_remove_dominant_line(
            segments,
            i.seg_indices,
            CONNECT_TOL * 0.9 * sc,
            SALVAGE_DIST_FACTOR,
            SALVAGE_LONG_SEG_FACTOR,
            SALVAGE_ANGLE_TOL_DEG,
            min_total=min_total,
        )
        for comp in subclusters:
            bx = _bbox_from_segments(segments, comp)
            w = bx[2] - bx[0]
            h = bx[3] - bx[1]
            tlen = sum(segments[j].length() for j in comp)
            if tlen < min_total:
                continue
            if (
                w > med_w * W_FACTOR
                or h > med_h * H_FACTOR
                or tlen > med_len * LEN_FACTOR
                or w * h > med_area * AREA_FACTOR
            ):
                continue
            if not aspect_ok(w, h):
                continue
            for split in _split_by_gap(
                segments,
                ClusterInfo(i.cluster_id, comp, bx, w, h, tlen, "digit_candidate"),
                med_w,
                med_h,
            ):
                final_infos.append(split)
    final_infos = [x for x in final_infos if is_renderable_cluster(segments, x)]
    final_infos = sorted(
        final_infos,
        key=lambda c: (-((c.bbox[1] + c.bbox[3]) / 2), (c.bbox[0] + c.bbox[2]) / 2),
    )
    return [
        Candidate(
            did,
            info.cluster_id,
            info.seg_indices,
            info.bbox,
            info.width,
            info.height,
            info.total_length,
        )
        for did, info in enumerate(final_infos)
    ]


def render_crop(segments, cand, out_size=128, pad_frac=0.15, thickness=2):
    minx, miny, maxx, maxy = cand.bbox
    w = maxx - minx
    h = maxy - miny
    if w < 1e-9 or h < 1e-9:
        return np.zeros((out_size, out_size), dtype=np.uint8)
    padx = pad_frac * w
    pady = pad_frac * h
    minx2, maxx2 = minx - padx, maxx + padx
    miny2, maxy2 = miny - pady, maxy + pady
    w2 = maxx2 - minx2
    h2 = maxy2 - miny2
    img = np.zeros((out_size, out_size), dtype=np.uint8)

    def to_px(x, y):
        return (
            int(round((x - minx2) / w2 * (out_size - 1))),
            int(round((1.0 - (y - miny2) / h2) * (out_size - 1))),
        )

    for si in cand.seg_indices:
        s = segments[si]
        if s.length() <= 1e-6:
            continue
        cv2.line(
            img,
            to_px(s.x1, s.y1),
            to_px(s.x2, s.y2),
            255,
            thickness=thickness,
            lineType=cv2.LINE_AA,
        )
    return img


def img_to_b64(img_np):
    _, buf = cv2.imencode(".png", img_np)
    return base64.b64encode(buf).decode()


# ─────────────────────────────────────────────────────────────────────────────
# CABLE SPAN HELPERS
# ─────────────────────────────────────────────────────────────────────────────


def find_cable_layer_names(layers: List[str], include_drops: bool = False) -> List[str]:
    """Layers holding the strand cable that the linemen will actually remove.

    Not everything named "Cable" is cable: Cable-840 marks power-supply
    locations on these drawings — it is a symbol, not plant to be torn down,
    and treating it as strand invented spans where no cable hangs. Excluded
    by default; override per deployment with CABLE_LAYER_EXCLUDE (comma-
    separated substrings, lowercase).

    SDU (subscriber drops) is teardown plant too, and every ---- of it shows
    in the whole-cable preview — but drops hang OFF the strand toward houses,
    they do not run pole to pole. Fed into the span graph they lace the
    streets together mid-block and the derivation explodes (5,752 "spans" on
    LP1709), so only the preview asks for them.
    """
    if not layers:
        return []
    matched = []

    # Add any substrings that identify your cable layers here (lowercase).
    keywords = ["cable", "tx56"]
    if include_drops:
        keywords = keywords + ["sdu"]
    exclude = [
        k.strip().lower()
        for k in os.environ.get("CABLE_LAYER_EXCLUDE", "840").split(",")
        if k.strip()
    ]

    for layer in layers:
        layer_lower = layer.lower()
        if not any(kw in layer_lower for kw in keywords):
            continue
        if any(ex in layer_lower for ex in exclude):
            continue
        matched.append(layer)

    return matched


def build_cable_spans(
    doc, cable_layers: List[str], connect_tol: float = CABLE_CONNECT_TOL
):
    """Build spans across MULTIPLE cable layers."""
    spans = []
    global_span_id = 0

    for cable_layer in cable_layers:
        segments = extract_stroke_segments(doc, cable_layer)
        if not segments:
            continue

        clusters = cluster_segments(segments, tol=connect_tol)

        for idxs in clusters:
            if not idxs:
                continue
            total_len = sum(segments[i].length() for i in idxs)
            if total_len <= 1e-8:
                continue

            bbox = _bbox_from_segments(segments, idxs)
            minx, miny, maxx, maxy = bbox
            cx = (minx + maxx) / 2.0
            cy = (miny + maxy) / 2.0

            span_segments = []
            for i in idxs:
                s = segments[i]
                if s.length() <= 1e-8:
                    continue
                span_segments.append(
                    {
                        "x1": s.x1,
                        "y1": s.y1,
                        "x2": s.x2,
                        "y2": s.y2,
                        "is_hatch": getattr(s, "is_hatch", False),
                    }
                )

            if not span_segments:
                continue

            # Store explicit cable endpoints so pole resolution can use proximity
            first_seg = span_segments[0]
            last_seg  = span_segments[-1]

            spans.append(
                {
                    "span_id": global_span_id,
                    "layer": cable_layer,
                    "bbox": [minx, miny, maxx, maxy],
                    "cx": cx,
                    "cy": cy,
                    "segment_count": len(span_segments),
                    "total_length": total_len,
                    "segments": span_segments,
                    "from_pole": None,
                    "to_pole": None,
                    # Explicit endpoints for proximity-based pole resolution
                    "from_x": first_seg["x1"],
                    "from_y": first_seg["y1"],
                    "to_x":   last_seg["x2"],
                    "to_y":   last_seg["y2"],
                }
            )
            global_span_id += 1

    spans.sort(key=lambda s: (-s["cy"], s["cx"]))
    for i, s in enumerate(spans):
        s["span_id"] = i

    return spans


def assign_meter_values_to_spans(spans, ocr_results, max_dist=None):
    if not spans:
        return spans
    if not ocr_results:
        for s in spans:
            s["meter_value"] = None
        return spans

    if max_dist is None:
        avg_size = sum(
            max(s["bbox"][2] - s["bbox"][0], s["bbox"][3] - s["bbox"][1]) for s in spans
        ) / len(spans)
        max_dist = avg_size * 0.5

    for span in spans:
        cx, cy = span["cx"], span["cy"]
        nearest_digit = None
        nearest_dist = float("inf")

        for r in ocr_results:
            dx = cx - r.get("center_x", 0)
            dy = cy - r.get("center_y", 0)
            dist = (dx**2 + dy**2) ** 0.5

            if dist < nearest_dist and dist <= max_dist:
                nearest_dist = dist
                nearest_digit = r

        if nearest_digit:
            value = nearest_digit.get("corrected_value") or nearest_digit.get("value")
            try:
                span["meter_value"] = float(value)
            except Exception:
                span["meter_value"] = None
        else:
            span["meter_value"] = None

    return spans


# ─────────────────────────────────────────────────────────────────────────────
# POLE-TOPOLOGY SPAN DERIVATION
#
# build_cable_spans() above answers "which cable segments are connected?" and
# calls each cluster a span, which is why broken linework exploded into many
# span ids and continuous linework hid the poles in between. span_builder
# answers "which poles are adjacent along the cable?" instead. Both live routes
# and both export paths read the result below, so they can no longer disagree
# about the same drawing.
# ─────────────────────────────────────────────────────────────────────────────

#: Last derivation, so exports never rebuild spans a second (differing) way.
SPAN_STATE: Dict = {"dxf_path": None, "result": None}


def derive_node_spans(
    dxf_path: str,
    poles: Optional[list] = None,
    ocr_results: Optional[list] = None,
) -> Tuple[Optional[span_builder.SpanBuildResult], List[str]]:
    """Derive pole-to-pole spans for a drawing, and cache the result.

    Poles and OCR values default to whatever the current session has detected.
    Returns ``(result, cable_layers)``; ``result`` is None only when the drawing
    has no cable layer at all.
    """
    doc = ezdxf.readfile(dxf_path)
    # Strand only — drops stay out of the span graph (see find_cable_layer_names).
    cable_layers = find_cable_layer_names(list_layers(dxf_path), include_drops=False)
    if not cable_layers:
        return None, []

    segments_by_layer = {
        layer: extract_dash_segments(doc, layer) for layer in cable_layers
    }
    # Streets drawn only on the *STRAND layers must form spans too, or the
    # poles along them stay unteardownable.
    base_pool = [
        s
        for segs in segments_by_layer.values()
        for s in segs
        if not getattr(s, "is_hatch", False) and s.length() > 1e-9
    ]
    supplemental = _supplemental_strand_segments(doc, dxf_path, base_pool)
    if supplemental:
        segments_by_layer.setdefault(cable_layers[0], []).extend(supplemental)
    poles = POLE_STATE.get("tags", []) if poles is None else poles
    ocr = state.get("results", []) if ocr_results is None else ocr_results

    result = span_builder.build_node_spans(segments_by_layer, poles, ocr)
    SPAN_STATE["dxf_path"] = dxf_path
    SPAN_STATE["result"] = result
    return result, cable_layers


def cached_span_result(dxf_path: str) -> Optional[span_builder.SpanBuildResult]:
    """The last derivation for this drawing, if it is still the current one."""
    if SPAN_STATE.get("dxf_path") == dxf_path:
        return SPAN_STATE.get("result")
    return None


def _length_weighted_median(walks: list) -> float:
    """The stroke length at which half the drawn INK is in shorter strokes.

    A plain median counts strokes, and a layer full of tiny text strokes drags
    it to text scale — which is how the SDU layer fell back to full extraction
    and poisoned every threshold downstream. Cable dashes dominate a layer by
    total length even when text outnumbers them.
    """
    if not walks:
        return 0.0
    ordered = sorted(walks)
    half = sum(ordered) / 2.0
    acc = 0.0
    for w in ordered:
        acc += w
        if acc >= half:
            return w
    return ordered[-1]


def extract_dash_segments(doc, layer_name: str) -> list:
    """The dashed cable linework of a layer, judged per drawn entity.

    Every ---- matters, and only the ----: a dash is drawn as its own short,
    straight polyline, while the curls beside poles, the text glyphs and the
    solid guy lines are entities of a different shape. Classifying whole
    entities is exact where welding tessellated segments back together was
    guesswork — this is what ended the noise-rule whack-a-mole.

    Falls back to the full extraction when a layer is not drafted as dashes
    (some drawings draw cable as continuous polylines), so span derivation
    still works on solid-drawn cable.
    """
    candidates = []  # (points, walk_len)
    other_len = 0.0
    for e in doc.modelspace():
        if getattr(e.dxf, "layer", None) != layer_name:
            continue
        t = e.dxftype()
        if t == "LINE":
            pts = [
                (float(e.dxf.start.x), float(e.dxf.start.y)),
                (float(e.dxf.end.x), float(e.dxf.end.y)),
            ]
        elif t == "LWPOLYLINE":
            pts = [(float(x), float(y)) for x, y, *_ in e.get_points("xy")]
        else:
            continue
        if len(pts) < 2:
            continue
        walk = sum(math.hypot(b[0] - a[0], b[1] - a[1]) for a, b in zip(pts, pts[1:]))
        if walk <= 1e-9:
            continue
        chord = math.hypot(pts[-1][0] - pts[0][0], pts[-1][1] - pts[0][1])
        if len(pts) <= 5 and chord / walk > 0.9:
            candidates.append((pts, walk))
        else:
            other_len += walk

    if not candidates:
        return extract_stroke_segments(doc, layer_name, include_circles=False)

    med = _length_weighted_median([w for _, w in candidates])
    # Drawings mix dash styles: the short-dash streets sit near the median,
    # the long-dash streets (WISDOM ran at 4-6x it) are still cable. Solid guy
    # lines start at ~19x, so the window stays well clear of them.
    dashes = [
        (pts, w) for pts, w in candidates if med * 0.3 <= w <= med * 8.0
    ]
    kept_len = sum(w for _, w in dashes)
    total_len = kept_len + other_len + sum(
        w for _, w in candidates if not (med * 0.3 <= w <= med * 3.0)
    )
    # A layer that is mostly NOT dashes was not drafted as dashed cable —
    # serve it whole rather than shredding it.
    if total_len > 0 and kept_len < 0.4 * total_len:
        return extract_stroke_segments(doc, layer_name, include_circles=False)

    segs = []
    for pts, _ in dashes:
        for a, b in zip(pts, pts[1:]):
            segs.append(Seg(a[0], a[1], b[0], b[1]))
    return segs


def _misfiled_street_trains(doc, dxf_path: str, base_pool: list) -> Dict[str, list]:
    """Street cable filed on the wrong layer, found by shape and connection.

    Candidates are dash-shaped entities (short, straight, dash-scale) on any
    layer not already treated as cable. They only count when they chain into a
    street-length train AND that train's end meets the known cable network —
    the two things lot fences, text and symbols never do together.
    """
    from app_python.services import span_builder as _sb

    # The dash-length reference comes from the cable layers' own dash ENTITIES
    # (length-weighted, like extract_dash_segments) — per-edge segment medians
    # sit at text scale and had every threshold below chasing letters.
    cable_walks: list = []
    cable_layer_names = set(
        find_cable_layer_names(list_layers(dxf_path), include_drops=True)
    )
    for e in doc.modelspace():
        if getattr(e.dxf, "layer", None) not in cable_layer_names:
            continue
        t = e.dxftype()
        if t == "LINE":
            pts = [
                (float(e.dxf.start.x), float(e.dxf.start.y)),
                (float(e.dxf.end.x), float(e.dxf.end.y)),
            ]
        elif t == "LWPOLYLINE":
            pts = [(float(x), float(y)) for x, y, *_ in e.get_points("xy")]
        else:
            continue
        if len(pts) < 2 or len(pts) > 5:
            continue
        walk = sum(math.hypot(b[0] - a[0], b[1] - a[1]) for a, b in zip(pts, pts[1:]))
        chord = math.hypot(pts[-1][0] - pts[0][0], pts[-1][1] - pts[0][1])
        if walk > 1e-9 and chord / walk > 0.9:
            cable_walks.append(walk)
    med = _length_weighted_median(cable_walks)
    if med <= 0:
        return {}

    skip = ("cable", "sdu", "tx56", "strand")
    cand_by_layer: Dict[str, list] = {}
    for e in doc.modelspace():
        layer = getattr(e.dxf, "layer", None)
        if not layer or any(k in layer.lower() for k in skip):
            continue
        t = e.dxftype()
        if t == "LINE":
            pts = [
                (float(e.dxf.start.x), float(e.dxf.start.y)),
                (float(e.dxf.end.x), float(e.dxf.end.y)),
            ]
        elif t == "LWPOLYLINE":
            pts = [(float(x), float(y)) for x, y, *_ in e.get_points("xy")]
        else:
            continue
        if len(pts) < 2 or len(pts) > 5:
            continue
        walk = sum(math.hypot(b[0] - a[0], b[1] - a[1]) for a, b in zip(pts, pts[1:]))
        if walk <= 1e-9:
            continue
        chord = math.hypot(pts[-1][0] - pts[0][0], pts[-1][1] - pts[0][1])
        if chord / walk < 0.9:
            continue
        if not (med * 0.4 <= walk <= med * 3.0):
            continue
        for a, b in zip(pts, pts[1:]):
            cand_by_layer.setdefault(layer, []).append(Seg(a[0], a[1], b[0], b[1]))

    # Grid over the known cable, to test train ends for connection.
    link_tol = med * 2.5
    cell = max(link_tol, 1e-9)
    grid: Dict[Tuple[int, int], list] = {}
    for s in base_pool:
        for gx, gy in ((s.x1, s.y1), (s.x2, s.y2)):
            grid.setdefault(
                (int(math.floor(gx / cell)), int(math.floor(gy / cell))), []
            ).append(s)

    def joins_network(p: Tuple[float, float]) -> bool:
        kx, ky = int(math.floor(p[0] / cell)), int(math.floor(p[1] / cell))
        for dx in (-1, 0, 1):
            for dy in (-1, 0, 1):
                for g in grid.get((kx + dx, ky + dy), ()):
                    d, _ = _sb._point_to_segment(p[0], p[1], g.x1, g.y1, g.x2, g.y2)
                    if d <= link_tol:
                        return True
        return False

    out: Dict[str, list] = {}
    min_train = med * 10
    for layer, segs in cand_by_layer.items():
        if len(segs) < 8:
            continue
        frags = _sb._build_fragments(segs, med * 0.2)
        trains = _sb.build_chains(frags, med * 3.0, med * 0.01)
        kept: list = []
        for tr in trains:
            if tr.total_length < min_train:
                continue
            # BOTH ends must meet the network: a misfiled street runs junction
            # to junction. A lot fence touches the street at one end and dies
            # into the block; a power-feed dead-ends at its supply; text
            # touches nothing. One-ended adoption let all three in.
            if joins_network(tr.points[0]) and joins_network(tr.points[-1]):
                for a, b in zip(tr.points, tr.points[1:]):
                    kept.append(Seg(a[0], a[1], b[0], b[1]))
        if kept:
            total = sum(s.length() for s in kept)
            print(f"[misfiled] {layer}: {len(kept)} seg(s), {total:.1f} units of street cable adopted")
            out[layer] = kept
    return out


def _supplemental_strand_segments(doc, dxf_path: str, base_pool: list) -> list:
    """Cable drawn on the STRAND layers where the Cable layers have nothing.

    Some streets carry their strand only on a *STRAND layer; ~88% of that layer
    parallels Cable-565 at a fixed small offset (the same cable drawn twice)
    and must not be doubled, but the rest is real plant on streets the Cable
    layers never cover — leaving it out left whole streets untraceable.
    """
    strand_layers = [
        l for l in list_layers(dxf_path) if "strand" in l.lower()
    ]
    if not strand_layers or not base_pool:
        return []

    lens = sorted(s.length() for s in base_pool if s.length() > 1e-9)
    med = lens[len(lens) // 2] if lens else 0.0
    if med <= 0:
        return []
    dup_tol = med * 3  # the co-drawn copy sits ~2x a stroke away; unique plant is farther

    cell = max(dup_tol * 2, 1e-9)
    grid: Dict[Tuple[int, int], list] = {}
    for s in base_pool:
        for gx, gy in ((s.x1, s.y1), (s.x2, s.y2)):
            grid.setdefault(
                (int(math.floor(gx / cell)), int(math.floor(gy / cell))), []
            ).append(s)

    from app_python.services import span_builder as _sb

    def near_base(mx: float, my: float) -> bool:
        kx, ky = int(math.floor(mx / cell)), int(math.floor(my / cell))
        for dx in (-1, 0, 1):
            for dy in (-1, 0, 1):
                for g in grid.get((kx + dx, ky + dy), ()):
                    d, _ = _sb._point_to_segment(mx, my, g.x1, g.y1, g.x2, g.y2)
                    if d <= dup_tol:
                        return True
        return False

    extra = []
    # Raw linework, not prepare_segments: the digit filter keys on the Cable
    # layers' colour split and eats most of a STRAND layer. The field rule is
    # that cable is the DASHED linework only. A dash is a short, STRAIGHT,
    # free-standing stroke; a strand-length digit is a tight knot of crooked
    # strokes, and its longer strokes (the tall side of a 4, say) slip past
    # any pure length cut — the trace was underlining numbers. So the strokes
    # are welded into fragments first and judged by shape: straight and simple
    # is cable, crooked or busy is a glyph. Solid guy/anchor lines fail the
    # length window as before.
    min_len = med * 0.5
    max_len = med * 2.5
    glyph_centres: List[Tuple[float, float]] = []
    candidates: List[Any] = []
    for layer in strand_layers:
        raw = [
            s
            for s in extract_stroke_segments(doc, layer, include_circles=False)
            if not getattr(s, "is_hatch", False) and s.length() > 1e-9
        ]
        if not raw:
            continue
        from app_python.services.span_builder import _build_fragments

        for frag in _build_fragments(raw, med * 0.2):
            cx = sum(p[0] for p in frag.points) / len(frag.points)
            cy = sum(p[1] for p in frag.points) / len(frag.points)
            span_dist = math.hypot(
                frag.points[-1][0] - frag.points[0][0],
                frag.points[-1][1] - frag.points[0][1],
            )
            # Crooked (walk much longer than end-to-end) or busy = a glyph.
            if len(frag.points) > 4 or span_dist < frag.length * 0.8:
                glyph_centres.append((cx, cy))
                continue
            if not (min_len <= frag.length <= max_len):
                if frag.length < min_len:
                    # Short strokes are almost always lettering too.
                    glyph_centres.append((cx, cy))
                continue
            if near_base(cx, cy):
                continue
            candidates.append((frag, cx, cy))

    # A single straight stroke can be a dash or the tall side of a 7 — the
    # length says nothing. Context does: digits come in knots beside other
    # glyphs, dashes come in trains of collinear neighbours.
    kept = 0
    for frag, cx, cy in candidates:
        if any(math.hypot(cx - gx, cy - gy) < med * 3 for gx, gy in glyph_centres):
            continue
        dx = frag.points[-1][0] - frag.points[0][0]
        dy = frag.points[-1][1] - frag.points[0][1]
        norm = math.hypot(dx, dy)
        if norm <= 1e-9:
            continue
        dx, dy = dx / norm, dy / norm
        in_train = False
        for other, ox, oy in candidates:
            if other is frag:
                continue
            d = math.hypot(ox - cx, oy - cy)
            if d > med * 4 or d <= 1e-9:
                continue
            odx = other.points[-1][0] - other.points[0][0]
            ody = other.points[-1][1] - other.points[0][1]
            onorm = math.hypot(odx, ody)
            if onorm <= 1e-9:
                continue
            if abs((odx * dx + ody * dy) / onorm) < 0.9:
                continue
            # The neighbour must sit along this dash's axis, not beside it.
            ux, uy = (ox - cx) / d, (oy - cy) / d
            if abs(ux * dx + uy * dy) < 0.9:
                continue
            in_train = True
            break
        if not in_train:
            continue
        kept += 1
        for a, b in zip(frag.points, frag.points[1:]):
            extra.append(Seg(a[0], a[1], b[0], b[1]))
    if glyph_centres or candidates:
        print(
            f"[strand] supplement: {kept} dash fragment(s) kept, "
            f"{len(candidates) - kept} lone/glyph-adjacent dropped, "
            f"{len(glyph_centres)} glyph fragment(s) fenced off"
        )
    if extra:
        print(f"[strand] {len(extra)} segment(s) taken from {strand_layers} where the Cable layers have no linework")
    return extra


def _simplify_polyline(
    pts: List[Tuple[float, float]], eps: float
) -> List[Tuple[float, float]]:
    """Douglas-Peucker: collapse jittery collinear stretches, keep corners."""
    if len(pts) < 3:
        return pts
    from app_python.services.span_builder import _point_to_segment

    ax, ay = pts[0]
    bx, by = pts[-1]
    worst_i, worst_d = 0, 0.0
    for i in range(1, len(pts) - 1):
        d, _ = _point_to_segment(pts[i][0], pts[i][1], ax, ay, bx, by)
        if d > worst_d:
            worst_i, worst_d = i, d
    if worst_d <= eps:
        return [pts[0], pts[-1]]
    left = _simplify_polyline(pts[: worst_i + 1], eps)
    right = _simplify_polyline(pts[worst_i:], eps)
    return left[:-1] + right


def _dash_polylines(pool: list, med: float) -> List[Dict[str, Any]]:
    """The dashed route as clean straight lines.

    Dashes are chained through their mutual-best joins into trains, and each
    train is simplified so the drafting jitter of individual dashes collapses
    into single straight strokes while corners survive. The trace follows the
    dashes exactly — it just reads as one drawn line, the way the cable
    actually hangs.
    """
    connectors = _dash_connectors(pool, med)

    # Rebuild the pairing the connectors encode: endpoint -> endpoint.
    def key(x: float, y: float) -> Tuple[int, int]:
        return (int(round(x * 1000)), int(round(y * 1000)))

    partner: Dict[Tuple[int, int], Tuple[float, float]] = {}
    for c in connectors:
        partner[key(c["x1"], c["y1"])] = (c["x2"], c["y2"])
        partner[key(c["x2"], c["y2"])] = (c["x1"], c["y1"])

    by_end: Dict[Tuple[int, int], list] = {}
    for s in pool:
        by_end.setdefault(key(s.x1, s.y1), []).append(s)
        by_end.setdefault(key(s.x2, s.y2), []).append(s)

    segs_out: List[Dict[str, Any]] = []
    visited: set = set()

    def walk(start_seg, start_pt: Tuple[float, float]):
        pts = [start_pt]
        train = []
        seg = start_seg
        entry = start_pt
        while True:
            visited.add(id(seg))
            train.append(seg)
            far = (
                (seg.x2, seg.y2)
                if key(*entry) == key(seg.x1, seg.y1)
                else (seg.x1, seg.y1)
            )
            pts.append(far)
            nxt_pt = partner.get(key(*far))
            if nxt_pt is None:
                break
            pts.append(nxt_pt)
            candidates = [s for s in by_end.get(key(*nxt_pt), []) if id(s) not in visited]
            if not candidates:
                break
            seg = candidates[0]
            entry = nxt_pt
        return pts, train

    from app_python.services.span_builder import _point_to_segment

    # Spatial index for the neighbour test below. It used to scan the whole
    # pool per lone dash — thousands of lone dashes on the SDU layer times a
    # 7,000-segment pool put ~35M distance checks in the request path, which
    # is where the two-minute viewer loads went.
    max_reach = max(max((s.length() for s in pool), default=0.0) * 2.2, med * 4)
    n_cell = max(max_reach, 1e-9)
    n_grid: Dict[Tuple[int, int], list] = {}
    for s in pool:
        n_grid.setdefault(
            (
                int(math.floor(((s.x1 + s.x2) / 2) / n_cell)),
                int(math.floor(((s.y1 + s.y2) / 2) / n_cell)),
            ),
            [],
        ).append(s)

    def _near_pool(cx: float, cy: float):
        kx, ky = int(math.floor(cx / n_cell)), int(math.floor(cy / n_cell))
        for dx2 in (-1, 0, 1):
            for dy2 in (-1, 0, 1):
                yield from n_grid.get((kx + dx2, ky + dy2), ())

    def has_collinear_neighbour(s, exclude: Optional[set] = None) -> bool:
        """A lone dash is cable only if another dash continues its axis."""
        cx, cy = (s.x1 + s.x2) / 2, (s.y1 + s.y2) / 2
        L = s.length()
        if L <= 1e-9:
            return False
        dx, dy = (s.x2 - s.x1) / L, (s.y2 - s.y1) / L
        for o in _near_pool(cx, cy):
            if o is s or (exclude and id(o) in exclude):
                continue
            ox, oy = (o.x1 + o.x2) / 2, (o.y1 + o.y2) / 2
            d = math.hypot(ox - cx, oy - cy)
            # Long-dash streets pitch at roughly their own dash length plus a
            # gap — a fixed med*4 radius, tuned to short dashes, called every
            # long dash "alone" and threw its street away.
            if d > max(med * 4, L * 2.2) or d <= 1e-9:
                continue
            oL = o.length()
            if oL <= 1e-9:
                continue
            odx, ody = (o.x2 - o.x1) / oL, (o.y2 - o.y1) / oL
            if abs(odx * dx + ody * dy) < 0.9:
                continue
            ux, uy = (ox - cx) / d, (oy - cy) / d
            if abs(ux * dx + uy * dy) < 0.85:
                continue
            return True
        return False

    eps = med * 0.6
    check_tol = med
    noise = 0
    kept_trains: List[Tuple[List[Tuple[float, float]], list, float]] = []
    for s in pool:
        if id(s) in visited:
            continue
        # Prefer starting at a free end so the walk covers the whole train.
        for start in ((s.x1, s.y1), (s.x2, s.y2)):
            if key(*start) not in partner:
                break
        pts, train = walk(s, start)

        # Short trains carry the noise: junction markers spray lone bars,
        # service loops curl into hooks, and stray symbol strokes form little
        # stubs of two or three dashes. Real cable this short is always going
        # somewhere — one of its ends continues into more cable along the same
        # axis. A short train that curls, or whose ends both dead-end, is a
        # drawn symbol and stays out.
        walk_len = sum(
            math.hypot(b[0] - a[0], b[1] - a[1]) for a, b in zip(pts, pts[1:])
        )
        span_dist = math.hypot(pts[-1][0] - pts[0][0], pts[-1][1] - pts[0][1])
        if walk_len < med * 6:
            if span_dist < walk_len * 0.7:
                noise += 1
                continue
            members = {id(t) for t in train}
            if not (
                has_collinear_neighbour(train[0], members)
                or has_collinear_neighbour(train[-1], members)
            ):
                noise += 1
                continue

        kept_trains.append((pts, train, walk_len))

    # Islands: a train that touches no other train anywhere is not part of the
    # network. Leader lines and callout strokes on the cable layer are long
    # enough to pass every per-train test, but real cable always joins the
    # grid somewhere — an unconnected train of modest length is drawing
    # furniture, not a street.
    link_tol = med * 6
    # Only genuinely small islands are drawing furniture. This started at
    # med*25 and twice swallowed real street segments (a 1.7-unit stretch of
    # WISDOM street among them) whose junction neighbours sat just past the
    # link tolerance.
    island_cap = med * 15
    # A train connects to the network wherever one of its ends meets ANY dash
    # of another train — including mid-street, which is exactly where a branch
    # road joins the one it feeds off.
    cell2 = max(link_tol, 1e-9)
    dash_grid: Dict[Tuple[int, int], List[Tuple[int, Any]]] = {}
    for idx, (_, train, _) in enumerate(kept_trains):
        for t in train:
            for gx, gy in ((t.x1, t.y1), (t.x2, t.y2)):
                dash_grid.setdefault(
                    (int(math.floor(gx / cell2)), int(math.floor(gy / cell2))), []
                ).append((idx, t))
    connected: set = set()
    for idx, (pts, _, _) in enumerate(kept_trains):
        for p in (pts[0], pts[-1]):
            kx, ky = int(math.floor(p[0] / cell2)), int(math.floor(p[1] / cell2))
            hit = False
            for dx in (-1, 0, 1):
                for dy in (-1, 0, 1):
                    for tj, t in dash_grid.get((kx + dx, ky + dy), ()):
                        if tj == idx:
                            continue
                        d, _u = _point_to_segment(p[0], p[1], t.x1, t.y1, t.x2, t.y2)
                        if d <= link_tol:
                            hit = True
                            break
                    if hit:
                        break
                if hit:
                    break
            if hit:
                connected.add(idx)
                break

    for idx, (pts, train, walk_len) in enumerate(kept_trains):
        if idx not in connected and walk_len < island_cap:
            noise += 1
            continue

        simplified = _simplify_polyline(pts, eps)

        # The clean line must actually lie on its dashes. A walk that went
        # wrong — a junction taken in the wrong order, a mismatched endpoint —
        # produces a stroke that departs from the route; that train falls back
        # to its raw dashes rather than showing cable where there is none.
        faithful = True
        for t in train:
            mx, my = (t.x1 + t.x2) / 2, (t.y1 + t.y2) / 2
            best = min(
                (
                    _point_to_segment(mx, my, a[0], a[1], b[0], b[1])[0]
                    for a, b in zip(simplified, simplified[1:])
                ),
                default=float("inf"),
            )
            if best > check_tol:
                faithful = False
                break

        if faithful:
            for a, b in zip(simplified, simplified[1:]):
                segs_out.append(
                    {
                        "x1": round(a[0], 6), "y1": round(a[1], 6),
                        "x2": round(b[0], 6), "y2": round(b[1], 6),
                    }
                )
        else:
            for t in train:
                segs_out.append(
                    {
                        "x1": round(t.x1, 6), "y1": round(t.y1, 6),
                        "x2": round(t.x2, 6), "y2": round(t.y2, 6),
                    }
                )
    if noise:
        print(f"[whole-cable] {noise} stray symbol stroke(s)/curl(s)/island(s) dropped from the preview")
    return segs_out


def _dash_connectors(pool: list, med: float) -> List[Dict[str, Any]]:
    """Short, direction-aligned joins between neighbouring dash ends.

    This is what makes the strand read as one line instead of dashes — without
    the chain walk, whose long bridges and junction zig-zags drew cable where
    the drawing has none.
    """
    from app_python.services import span_builder as _sb

    # The allowed join gap follows the dash it joins: a long-dash street's
    # gaps run near its own dash length, far past the short-dash median that
    # med*3 was tuned to — those streets never chained and then died as
    # "lone dead-end dashes".
    max_len = max((s.length() for s in pool), default=med)
    max_gap_global = max(med * 3, max_len * 0.9)
    cell = max(max_gap_global, 1e-9)
    ends: List[Tuple[float, float, float, float, float]] = []  # x, y, dirx, diry, dash_len
    for s in pool:
        L = s.length()
        if L <= 1e-9:
            continue
        dx, dy = (s.x2 - s.x1) / L, (s.y2 - s.y1) / L
        ends.append((s.x1, s.y1, -dx, -dy, L))
        ends.append((s.x2, s.y2, dx, dy, L))

    grid: Dict[Tuple[int, int], List[int]] = {}
    for i, (x, y, _, _, _l) in enumerate(ends):
        grid.setdefault((int(math.floor(x / cell)), int(math.floor(y / cell))), []).append(i)

    def best_partner(i: int) -> Optional[int]:
        x, y, dx, dy, li = ends[i]
        kx, ky = int(math.floor(x / cell)), int(math.floor(y / cell))
        best_j, best_d = None, max_gap_global
        for gx in (-1, 0, 1):
            for gy in (-1, 0, 1):
                for j in grid.get((kx + gx, ky + gy), ()):
                    if j // 2 == i // 2:
                        continue
                    jx, jy, jdx, jdy, lj = ends[j]
                    d = math.hypot(jx - x, jy - y)
                    # Per-pair allowance: two long dashes may join across a
                    # gap near their own length; short dashes stay tight.
                    allowed = max(med * 3, min(li, lj) * 0.9)
                    if d >= min(best_d, allowed) or d <= 1e-9:
                        continue
                    # The joint must continue this dash's direction and meet
                    # the other dash roughly head-on.
                    ux, uy = (jx - x) / d, (jy - y) / d
                    if ux * dx + uy * dy < 0.85:
                        continue
                    if ux * jdx + uy * jdy > -0.85:
                        continue
                    best_j, best_d = j, d
        return best_j

    best = [best_partner(i) for i in range(len(ends))]

    connectors: List[Dict[str, Any]] = []
    for i, j in enumerate(best):
        # Mutual choice only. One-sided joins are how curves grew triangles:
        # the two ends across a bend could each still "see" the far side within
        # tolerance and a chord cut the corner, on top of the two real joins.
        # A chord can never beat the true next dash for BOTH of its ends.
        if j is None or j <= i:
            continue
        if best[j] != i:
            continue
        x, y, _, _, _li = ends[i]
        jx, jy, _, _, _lj = ends[j]
        connectors.append(
            {
                "x1": round(x, 6), "y1": round(y, 6),
                "x2": round(jx, 6), "y2": round(jy, 6),
            }
        )
    return connectors


#: dxf_path -> (file mtime, spans, cable_layers). The whole-cable preview is a
#: pure function of the file, so it is computed once per drawing, not once per
#: page load — it was costing two minutes on every viewer mount.
_WHOLE_CABLE_CACHE: Dict[str, Tuple[float, list, list]] = {}

#: (dxf_path, hide_circles) -> (file mtime, serialized JSON body).
_DXF_SEGMENTS_CACHE: Dict[Tuple[str, bool], Tuple[float, str]] = {}


def _whole_cable_spans(dxf_path: str) -> Tuple[List[Dict[str, Any]], List[str]]:
    """One span per cable layer — the whole strand, undivided.

    Before the pole scan there is nothing to cut the cable at, but the operator
    still needs to point at it: clicking anywhere on the strand selects it as
    one whole, and any toggle applies to all of it. The pole scan then replaces
    these with real pole-to-pole spans.
    """
    try:
        mtime = Path(dxf_path).stat().st_mtime
    except OSError:
        mtime = 0.0
    cached = _WHOLE_CABLE_CACHE.get(dxf_path)
    if cached and cached[0] == mtime:
        return cached[1], cached[2]

    doc = ezdxf.readfile(dxf_path)
    # Every ---- shows here, drops included — the preview is the inventory eye.
    cable_layers = find_cable_layer_names(list_layers(dxf_path), include_drops=True)
    spans: List[Dict[str, Any]] = []
    base_pool: list = []
    per_layer: Dict[str, list] = {}
    for layer in cable_layers:
        raw = extract_dash_segments(doc, layer)
        pool = span_builder.prepare_segments({layer: raw}, [])
        if pool:
            per_layer[layer] = pool
            base_pool.extend(pool)

    # Streets whose strand lives only on the *STRAND layers join the trace too.
    supplemental = _supplemental_strand_segments(doc, dxf_path, base_pool)
    if supplemental and per_layer:
        first = next(iter(per_layer))
        per_layer[first] = per_layer[first] + supplemental

    # Every ---- matters, and drafters file street cable on whatever layer was
    # active — LP1709 keeps whole streets on PDF_0 and Actives-PowerSupply.
    # A dashed train from any other layer joins the preview when it is street-
    # length AND its end meets the cable network — lot fences and text never
    # connect end-on to the strand, misfiled streets always do.
    # NOTE: _misfiled_street_trains stays available but is not called — the
    # "missing streets" it was written for turned out to be long-dash Cable-565
    # linework (handled by the adaptive dash rules), while its own catches on
    # LP1709 were road-name text and lot fences.

    for layer, pool in per_layer.items():
        # Every drawn dash, exactly as drafted, plus short direction-aligned
        # joins between dash ends. The earlier chain-walk render invented
        # cable — long bridges and junction zig-zags where the drawing has
        # nothing — and still dropped branch sides; dashes cannot lie.
        med = span_builder._median([s.length() for s in pool])
        segs = _dash_polylines(pool, med)
        if not segs:
            continue
        xs = [v for g in segs for v in (g["x1"], g["x2"])]
        ys = [v for g in segs for v in (g["y1"], g["y2"])]
        bbox = [min(xs), min(ys), max(xs), max(ys)]
        total = sum(s.length() for s in pool)
        spans.append(
            {
                "span_id": len(spans),
                "span_key": None,
                "whole_cable": True,
                # The viewer strokes the preview with this dash pattern so it
                # reads like the drawing's own ----, not a solid bar.
                "dash_len": round(med, 6),
                "layer": layer,
                "segments": segs,
                "segment_count": len(segs),
                "bbox": [round(v, 6) for v in bbox],
                "cx": round((bbox[0] + bbox[2]) / 2, 6),
                "cy": round((bbox[1] + bbox[3]) / 2, 6),
                "total_length": round(total, 4),
                "arc_length": round(total, 4),
                "strand_length": round(total, 4),
                "meter_value": None,
                "length_source": "arc_length",
                "cable_runs": 1,
                "from_pole": None,
                "to_pole": None,
                "from_pole_id": None,
                "to_pole_id": None,
            }
        )
    _WHOLE_CABLE_CACHE[dxf_path] = (mtime, spans, cable_layers)
    return spans, cable_layers


def _span_response(dxf_path: str) -> Dict[str, Any]:
    """Shared body for both cable_spans routes.

    Poles are detected by a separate job the client kicks off, so a drawing can
    legitimately have no poles yet. That is a waiting state, not an error — the
    viewer polls this route from the moment it mounts, and until poles exist it
    gets the whole strand as one selectable span per layer.
    """
    poles = POLE_STATE.get("tags", [])
    if not poles:
        spans, cable_layers = _whole_cable_spans(dxf_path)
        return {
            "cable_layers": cable_layers,
            "count": len(spans),
            "spans": spans,
            "poles": [],
            "warnings": [],
            "errors": [],
            "status": "awaiting_poles",
        }

    result, cable_layers = derive_node_spans(dxf_path)
    if result is None:
        return {
            "cable_layers": [],
            "count": 0,
            "spans": [],
            "poles": [],
            "warnings": [],
            "errors": [],
            "status": "no_cable_layer",
            "message": "No cable layers found",
        }

    body = result.to_dict()
    body["cable_layers"] = cable_layers
    body["status"] = "ok" if result.ok else "blocked"
    return body


# ─────────────────────────────────────────────────────────────────────────────
# PLANNER INTEGRATION
# ─────────────────────────────────────────────────────────────────────────────


def _pole_id_key(value) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _safe_code_part(value, fallback: str = "X") -> str:
    text = str(value or "").strip().upper()
    text = _re.sub(r"[^A-Z0-9]+", "-", text).strip("-")
    return text or fallback


def push_to_planner(
    dxf_path: str,
    poles: list,
    spans: list,
    equipment: list,
    ocr_results: list,
    project_id: int,
) -> dict:
    if not ENABLE_PLANNER_INTEGRATION:
        print("[planner] Integration disabled, skipping push.")
        return {"skipped": True, "reason": "Planner integration disabled"}

    try:
        print(f"[planner] Starting bulk push for {dxf_path}")

        dxf_name = Path(dxf_path).stem if dxf_path else "CAD_NODE"
        node_id = dxf_name.split("_")[0] if "_" in dxf_name else dxf_name
        if not node_id:
            node_id = "CAD_NODE"
        node_code = _safe_code_part(node_id, "CAD")

        equipment_counts = {"amplifier": 0, "extender": 0, "tsc": 0}
        for shape in equipment:
            kind = shape.get("kind", "")
            if kind in ("circle", "square", "hexagon"):
                equipment_counts["tsc"] += 1
            elif kind == "rectangle":
                equipment_counts["amplifier"] += 1
            elif kind == "triangle":
                equipment_counts["extender"] += 1

        spans = assign_meter_values_to_spans(spans, ocr_results)

        poles_list = []
        pole_id_map: dict[str, str] = {}
        pole_counter = 1

        # Detect pole names that appear more than once (NPT, PT, NT, etc.)
        # These get an index suffix so each is uniquely identifiable.
        _raw_name_counts: Counter = Counter(
            (p.get("corrected_name") or p.get("name", "")).strip().upper()
            for p in poles
            if (p.get("corrected_name") or p.get("name", "")).strip()
        )
        _duplicate_names = {n for n, c in _raw_name_counts.items() if c > 1}
        _dup_idx: dict[str, int] = {}

        for pole in poles:
            base_name = (pole.get("corrected_name") or pole.get("name", "")).strip()
            if not base_name:
                continue

            base_upper = base_name.upper()

            # Assign index to duplicate-named poles: NPT → NPT-1, NPT-2 …
            if base_upper in _duplicate_names:
                _dup_idx[base_upper] = _dup_idx.get(base_upper, 0) + 1
                pole_name = f"{base_name}-{_dup_idx[base_upper]}"
            else:
                pole_name = base_name

            pid = pole.get("pole_id")
            pid_key = _pole_id_key(pid)
            pole_code_suffix = _safe_code_part(pid_key, f"{pole_counter:03d}")
            pole_code = f"{node_code}-P{pole_code_suffix}"
            pole_counter += 1

            if pid_key is not None:
                pole_id_map[pid_key] = pole_code

            poles_list.append(
                {
                    "pole_code":    pole_code,
                    "pole_name":    pole_name,      # e.g. "NPT-1"
                    "_base_name":   base_upper,     # e.g. "NPT"  (for reverse lookup)
                    "pole_id":      pid,
                    "map_latitude": pole.get("cy"),
                    "map_longitude":pole.get("cx"),
                }
            )

        print(f"[planner] Built {len(poles_list)} poles for upload")
        print(f"[planner] Indexed duplicate names: {sorted(_duplicate_names)}")
        print(f"[planner] pole_id_map keys: {sorted(pole_id_map.keys())}")

        # All pole candidates for position-based fallback (CAD coords)
        all_pole_candidates = [
            p for p in poles_list
            if p["map_longitude"] is not None and p["map_latitude"] is not None
        ]

        # Fallback name→code map for spans that don't have pole_id.
        # Register each pole under BOTH its indexed name ("NPT-1") AND its base
        # name ("NPT") so spans that only say "NPT" can still resolve by proximity.
        name_code_map: dict[str, str] = {}
        name_candidates: dict[str, list] = defaultdict(list)
        for p in poles_list:
            indexed_key = (p["pole_name"] or "").strip().upper()
            base_key    = p.get("_base_name", indexed_key)

            if indexed_key:
                name_candidates[indexed_key].append(p)
                name_code_map.setdefault(indexed_key, p["pole_code"])

            # Also register under base name so "NPT" resolves to [NPT-1, NPT-2, ...]
            if base_key and base_key != indexed_key:
                name_candidates[base_key].append(p)
                name_code_map.setdefault(base_key, p["pole_code"])

        def _as_float(value):
            try:
                if value is None:
                    return None
                return float(value)
            except (TypeError, ValueError):
                return None

        def _span_endpoint(span, side: str):
            # Priority 1: explicit endpoint stored by build_cable_spans
            x = _as_float(span.get(f"{side}_x"))
            y = _as_float(span.get(f"{side}_y"))
            if x is not None and y is not None:
                return x, y

            # Priority 2: legacy field names
            x = _as_float(span.get(f"{side}_pole_x"))
            y = _as_float(span.get(f"{side}_pole_y"))
            if x is not None and y is not None:
                return x, y

            # Priority 3: derive from segments
            segments = span.get("segments") or []
            if not segments:
                return None

            if side == "from":
                seg = segments[0]
                x = _as_float(seg.get("x1"))
                y = _as_float(seg.get("y1"))
            else:
                seg = segments[-1]
                x = _as_float(seg.get("x2"))
                y = _as_float(seg.get("y2"))

            if x is None or y is None:
                return None
            return x, y

        def _nearest_candidate_match(candidates, endpoint):
            if endpoint is None:
                return None, None

            ex, ey = endpoint
            best_candidate = None
            best_dist = float("inf")
            for candidate in candidates:
                px = _as_float(candidate.get("map_longitude"))
                py = _as_float(candidate.get("map_latitude"))
                if px is None or py is None:
                    continue

                dist = math.hypot(px - ex, py - ey)
                if dist < best_dist:
                    best_dist = dist
                    best_candidate = candidate

            return best_candidate, best_dist if best_candidate is not None else None

        def _resolve_span_pole_code(span, side: str, span_idx: int = -1):
            pole_id = span.get(f"{side}_pole_id")
            pole_key = _pole_id_key(pole_id)
            if pole_key is not None:
                code = pole_id_map.get(pole_key)
                if code:
                    return code, "pole_id"

            # Log when pole_id lookup fails
            if span_idx >= 0 and pole_key is not None:
                print(f"[planner] Span #{span_idx} {side}_pole_id={pole_id!r} NOT found in pole_id_map (keys={sorted(pole_id_map.keys())})")

            pole_name = (span.get(f"{side}_pole") or "").strip().upper()
            if not pole_name:
                return None, "missing_name"

            candidates = name_candidates.get(pole_name, [])

            # --- Exact unique match ---
            if len(candidates) == 1:
                return candidates[0].get("pole_code"), "unique_name"

            # --- Multiple candidates (duplicate names like NPT-1/NPT-2 or raw "NPT") ---
            # Always try proximity first using the span's explicit cable endpoints.
            if len(candidates) > 1:
                endpoint = _span_endpoint(span, side)
                nearest_candidate, nearest_dist = _nearest_candidate_match(candidates, endpoint)
                if nearest_candidate:
                    return (
                        nearest_candidate.get("pole_code"),
                        f"duplicate_name_nearest:{nearest_dist:.4f}",
                    )
                # Proximity failed (no position data on candidates) →
                # widen search to all positioned poles.
                if span_idx >= 0:
                    print(
                        f"[planner] Span #{span_idx} {side}: {pole_name!r} has "
                        f"{len(candidates)} candidates but no usable endpoint; "
                        f"falling back to all-poles proximity"
                    )
                nearest_any, dist_any = _nearest_candidate_match(all_pole_candidates, endpoint)
                if nearest_any:
                    return nearest_any.get("pole_code"), f"duplicate_fallback_any:{dist_any:.4f}"
                # Last resort: just pick the first candidate rather than dropping the span
                return candidates[0].get("pole_code"), "duplicate_first_fallback"

            # --- No candidates found by name at all ---
            # Use proximity against all positioned poles (handles NPT/PT/NT
            # that appear in span data but whose indexed variant isn't looked up).
            endpoint = _span_endpoint(span, side)
            nearest_candidate, nearest_dist = _nearest_candidate_match(
                all_pole_candidates, endpoint
            )
            if nearest_candidate:
                if span_idx >= 0:
                    print(
                        f"[planner] Span #{span_idx} {side}: no name match for "
                        f"{pole_name!r}; position fallback → "
                        f"{nearest_candidate.get('pole_code')} (dist={nearest_dist:.4f})"
                    )
                return (
                    nearest_candidate.get("pole_code"),
                    f"nearest_any:{nearest_dist:.4f}",
                )

            # Unique non-NPT labels can still use the original name fallback.
            code = name_code_map.get(pole_name)
            return (code, "name_fallback") if code else (None, "no_candidate")

        pole_spans = []
        pole_pair_counts = {}  # pole pair -> the one span entry for it
        spans_no_pole_name = 0
        spans_no_code = 0
        spans_same_code = 0
        spans_ok = 0
        resolution_methods = Counter()
        span_resolution_debug = []

        for span_idx, span in enumerate(spans):
            from_pole = span.get("from_pole")
            to_pole = span.get("to_pole")

            if not from_pole or not to_pole:
                spans_no_pole_name += 1
                continue
            from_code, from_method = _resolve_span_pole_code(span, "from", span_idx)
            to_code, to_method = _resolve_span_pole_code(span, "to", span_idx)
            resolution_methods[f"from:{from_method}"] += 1
            resolution_methods[f"to:{to_method}"] += 1

            if len(span_resolution_debug) < 25:
                span_resolution_debug.append(
                    {
                        "span_id": span.get("span_id", span_idx),
                        "from_pole": from_pole,
                        "from_pole_id": span.get("from_pole_id"),
                        "from_code": from_code,
                        "from_method": from_method,
                        "to_pole": to_pole,
                        "to_pole_id": span.get("to_pole_id"),
                        "to_code": to_code,
                        "to_method": to_method,
                    }
                )

            if not from_code or not to_code:
                spans_no_code += 1
                continue
            if from_code == to_code:
                spans_same_code += 1
                continue
            spans_ok += 1

            # Two entries on one pole pair mean one physical span, so their
            # cable adds up. The old code appended -2, -3, -4 instead, which is
            # where the duplicate span ids in the backend came from: a lineman
            # tore the span down once and the extra ids stayed pending forever,
            # so the pole could never be cleared.
            pole_pair = tuple(sorted([from_code, to_code]))
            length = span.get("meter_value") or span.get("total_length", 0) or 0
            existing = pole_pair_counts.get(pole_pair)
            if existing is not None:
                existing["length_meters"] += length
                existing["expected_cable"] += length
                existing["runs"] = max(existing["runs"], span.get("cable_runs", 1))
                continue

            entry = {
                "from_pole_code": from_code,
                "to_pole_code": to_code,
                "pole_span_code": f"{node_id}-{from_code}-{to_code}",
                "length_meters": length,
                "runs": span.get("cable_runs", 1),
                "expected_cable": length,
            }
            pole_pair_counts[pole_pair] = entry
            pole_spans.append(entry)

        print(
            f"[planner] Built {len(pole_spans)} pole spans for upload "
            f"(skip reasons: {spans_no_pole_name} no-pole-name, {spans_no_code} no-code, {spans_same_code} same-code)"
        )

        total_strand_length = sum(s.get("length_meters", 0) for s in pole_spans)
        node_data = {
            "node_id": node_id,
            "node_name": node_id,
            "total_strand_length": total_strand_length,
            "expected_cable": total_strand_length,
            "node_count": 1,
            "date_start": datetime.now().strftime("%Y-%m-%d"),
            **equipment_counts,
        }

        payload = {
            "project_id": project_id,
            "node": node_data,
            "poles": poles_list,
            "pole_spans": pole_spans,
        }

        debug_dir = Path("cache")
        debug_dir.mkdir(exist_ok=True)
        debug_path = debug_dir / "polemaster-last-payload.json"
        debug_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
        print(f"[planner] Saved last bulk payload: {debug_path.resolve()}")

        print(
            f"[planner] Uploading bulk payload: {len(poles_list)} poles, {len(pole_spans)} spans"
        )

        result = auth.bulk_upload(payload)

        print(f"[planner] Bulk upload successful: {result.get('message', 'OK')}")

        data = result.get("data", result)
        summary = data.get("summary", {})
        spans_skipped = spans_no_pole_name + spans_no_code + spans_same_code
        duplicate_pole_names = {
            name: len(candidates)
            for name, candidates in name_candidates.items()
            if len(candidates) > 1
        }

        return {
            "success": True,
            "node_id": data.get("node", {}).get("id"),
            "node_action": data.get("node", {}).get("action", "created"),
            "poles_created": summary.get("poles_count", len(poles_list)),
            "spans_created": summary.get("pole_spans_count", len(pole_spans)),
            "spans_ready": spans_ok,
            "spans_skipped": spans_skipped,
            "spans_skipped_no_poles": spans_no_pole_name,
            "spans_skipped_unresolved": spans_no_code,
            "spans_skipped_same_pole": spans_same_code,
            "span_resolution_methods": dict(resolution_methods),
            "duplicate_pole_names": duplicate_pole_names,
            "span_resolution_sample": span_resolution_debug,
            "debug_payload_path": str(debug_path.resolve()),
        }

    except Exception as e:
        print(f"[planner] Bulk push failed: {e}")
        import traceback

        traceback.print_exc()
        return {"success": False, "error": str(e)}


# ─────────────────────────────────────────────────────────────────────────────
# GLOBAL STATE  (OCR pipeline)
# ─────────────────────────────────────────────────────────────────────────────

state = {
    "dxf_path": None,
    "model_path": None,
    "layers": [],
    "segments": [],
    "candidates": [],
    "results": [],
    "status": "idle",
    "progress": 0,
    "total": 0,
    "error": None,
    "step": 0,
    "step_label": "",
    "ocr_start_time": None,
}


def _clear_all_states():
    for k in list(state.keys()):
        state[k] = None
    state["layers"] = []
    state["segments"] = []
    state["candidates"] = []
    state["results"] = []
    state["status"] = "idle"
    state["progress"] = 0
    state["total"] = 0
    state["error"] = None
    state["step"] = 0
    state["step_label"] = ""
    state["ocr_start_time"] = None


# ─────────────────────────────────────────────────────────────────────────────
# POST-OCR VALIDATION
# ─────────────────────────────────────────────────────────────────────────────


def _post_ocr_validate(results: list) -> list:
    from collections import Counter

    value_counts = Counter(
        r.get("corrected_value") or r.get("value", "") for r in results
    )

    flagged = 0
    for r in results:
        if r.get("needs_review"):
            continue

        val = r.get("corrected_value") or r.get("value", "")
        conf = r.get("confidence", 0.0)
        freq = value_counts.get(val, 0)

        is_single_digit = val.isdigit() and len(val) == 1 and val != "0"
        if is_single_digit and freq < 3:
            r["needs_review"] = True
            flagged += 1
            continue

        if freq == 1 and conf < 0.80:
            r["needs_review"] = True
            flagged += 1
            continue

        if is_single_digit and conf < 0.90:
            r["needs_review"] = True
            flagged += 1

    return results


# ─────────────────────────────────────────────────────────────────────────────
# OCR PIPELINE
# ─────────────────────────────────────────────────────────────────────────────


def run_pipeline(dxf_path, layers, model_path):
    try:
        state.update(
            {
                "status": "processing",
                "progress": 0,
                "total": 0,
                "error": None,
                "step": 1,
                "step_label": "Extracting stroke segments…",
                "ocr_start_time": None,
                "results": [],
            }
        )

        doc = ezdxf.readfile(dxf_path)
        all_segments = []
        for lyr in layers:
            segs = extract_stroke_segments(doc, lyr, include_circles=False)
            all_segments.extend(segs)

        state["segments"] = all_segments

        state.update({"step": 2, "step_label": "Grouping into digit clusters…"})
        # Derive all distance thresholds from THIS drawing's stroke scale so the
        # pipeline works across maps of different sizes without hand-tuning.
        scale = estimate_scale(all_segments)
        # Exclude the cable strand line from clustering so it can't swallow
        # digits or leak a stub into a digit crop.
        cable = cable_segment_indices(all_segments, scale=scale)
        print(f"[run_pipeline] adaptive scale={scale:.3f}  "
              f"connect_tol={CONNECT_TOL * scale:.4f}  cable_segs={len(cable)}  "
              f"(ref={REF_MEDIAN_SEGLEN})")
        clusters = cluster_segments(all_segments, tol=CONNECT_TOL * scale, ignore=cable)
        infos = analyze_clusters(all_segments, clusters, scale=scale)

        state.update({"step": 3, "step_label": "Identifying digit candidates…"})
        candidates = build_candidates_robust(all_segments, infos, scale=scale)
        state["candidates"] = candidates
        state["total"] = len(candidates)

        if not candidates:
            state.update(
                {
                    "status": "done",
                    "step": 4,
                    "step_label": "Done — no candidates found",
                }
            )
            return

        crops = [render_crop(all_segments, cand) for cand in candidates]

        state.update(
            {
                "step": 4,
                "step_label": f"Reading digit 0 of {len(candidates)}…",
                "ocr_start_time": time.time(),
            }
        )

        # Recognize the whole drawing in one batch so flip ambiguities (e.g.
        # 19 vs 61) are resolved against the drawing-wide dominant orientation.
        from app_python.services.strand_recognizer import recognize_batch

        def _progress(done: int, total: int) -> None:
            elapsed = time.time() - state["ocr_start_time"]
            rate = done / elapsed if elapsed > 0 else 0
            eta_secs = int((total - done) / rate) if rate > 0 else 0
            if eta_secs >= 60:
                eta_str = f"{eta_secs // 60}m {eta_secs % 60}s remaining"
            elif eta_secs > 0:
                eta_str = f"~{eta_secs}s remaining"
            else:
                eta_str = "almost done…"
            state.update(
                {
                    "progress": done,
                    "step_label": f"Reading digit {done} of {total} — {eta_str}",
                }
            )

        batch = recognize_batch(crops, progress_cb=_progress)

        results = []
        for cand, crop, (value, conf) in zip(candidates, crops, batch):
            cx = (cand.bbox[0] + cand.bbox[2]) / 2
            cy = (cand.bbox[1] + cand.bbox[3]) / 2
            needs_review = (not value) or (conf < _MIN_CONF)
            results.append(
                {
                    "digit_id": cand.digit_id,
                    "value": value if value else "?",
                    "corrected_value": None,
                    "confidence": round(conf, 4),
                    "needs_review": needs_review,
                    "bbox": list(cand.bbox),
                    "center_x": cx,
                    "center_y": cy,
                    "crop_b64": img_to_b64(crop),
                }
            )

        results = _post_ocr_validate(results)

        state.update(
            {
                "status": "done",
                "step": 4,
                "step_label": "Done",
                "results": results,
            }
        )

    except Exception as e:
        import traceback

        traceback.print_exc()
        state.update({"status": "error", "error": str(e)})


# ─────────────────────────────────────────────────────────────────────────────
# SHAPE / EQUIPMENT DETECTION
# ─────────────────────────────────────────────────────────────────────────────

from app_python.services.boundary_service import (
    apply_boundary_filter,
    build_boundary_mask,
)
from app_python.services.shape_service import extract_equipment_shapes

SCAN_STATE = {
    "status": "idle",
    "error": None,
    "shapes": [],
    "boundary": None,
    "progress": 0,
    "total": 0,
}

SHAPE_CONFIG = {
    "min_circle_r": 1e-5,
    "min_poly_area": 1e-6,
    "dedup_eps": 1e-4,
    "min_rect_short_side": 0.05,
    "max_rect_aspect": 50.0,
}

BOUNDARY_CONFIG = {
    "snap_tol": 0.60,
    "close_max_gap": 2.50,
    "min_area": 1e-6,
}


def _run_full_scan(dxf_path: str, boundary_layer: Optional[str]):
    try:
        SCAN_STATE.update(
            {
                "status": "processing",
                "error": None,
                "shapes": [],
                "boundary": None,
                "progress": 0,
                "total": 0,
            }
        )

        doc = ezdxf.readfile(dxf_path)
        layers = list_layers(dxf_path)

        KIND_LAYER_MAP = {
            "circle": ["splitter", "tapoff", "tap-off", "tap_off", "splt"],
            "hexagon": ["tapoff", "tap-off", "tap_off"],
            "rectangle": ["node", "amplifier", "amp"],
            "square": ["tapoff", "tap-off", "tap_off"],
            "triangle": ["extender", "extend"],
        }

        layer_kind_targets: Dict[str, List[str]] = {}
        for layer in layers:
            if boundary_layer and layer == boundary_layer:
                continue
            l_lower = layer.lower()
            kinds_for_layer = []
            for kind, keywords in KIND_LAYER_MAP.items():
                if any(kw in l_lower for kw in keywords):
                    kinds_for_layer.append(kind)
            if kinds_for_layer:
                layer_kind_targets[layer] = kinds_for_layer

        scan_layers = list(layer_kind_targets.keys())
        SCAN_STATE["total"] = len(scan_layers)

        all_shapes = []
        for i, layer in enumerate(scan_layers):
            allowed_kinds = set(layer_kind_targets[layer])
            try:
                shapes = extract_equipment_shapes(doc, layer, **SHAPE_CONFIG)
                for s in shapes:
                    if s.kind not in allowed_kinds:
                        continue
                    all_shapes.append(
                        {
                            "shape_id": -1,
                            "kind": s.kind,
                            "bbox": list(s.bbox),
                            "cx": s.cx,
                            "cy": s.cy,
                            "layer": layer,
                        }
                    )
            except Exception as e:
                pass
            SCAN_STATE["progress"] = i + 1

        DEDUP_EPS = 0.5
        all_shapes.sort(key=lambda s: (s["kind"], s["cx"], s["cy"]))
        deduped = []
        for s in all_shapes:
            if not deduped:
                deduped.append(s)
                continue
            prev = deduped[-1]
            if (
                s["kind"] == prev["kind"]
                and abs(s["cx"] - prev["cx"]) < DEDUP_EPS
                and abs(s["cy"] - prev["cy"]) < DEDUP_EPS
            ):
                continue
            deduped.append(s)

        deduped.sort(key=lambda s: (-s["cy"], s["cx"]))
        for i, s in enumerate(deduped):
            s["shape_id"] = i

        boundary_pts = None
        if boundary_layer:
            try:
                boundary = build_boundary_mask(
                    doc, boundary_layer=boundary_layer, **BOUNDARY_CONFIG
                )
                if boundary:
                    boundary_pts = [{"x": p[0], "y": p[1]} for p in boundary.pts]
            except Exception as e:
                pass

        highleg_layer = next((l for l in layers if "highleg" in l.lower()), None)
        if highleg_layer:
            highleg_segs = extract_stroke_segments(doc, highleg_layer)
            if highleg_segs:
                for s in deduped:
                    # Target circles strictly on the splt layer
                    if s["kind"] == "circle" and "splt" in s.get("layer", "").lower():
                        w = s["bbox"][2] - s["bbox"][0]
                        h = s["bbox"][3] - s["bbox"][1]
                        r = max(w, h) * 0.5

                        # Set threshold to 3x the radius to account for the gap in the drawing
                        threshold = r * 3.0

                        is_splitter = False
                        for seg in highleg_segs:
                            d = point_segment_dist(
                                s["cx"], s["cy"], seg.x1, seg.y1, seg.x2, seg.y2
                            )
                            if d <= threshold:
                                is_splitter = True
                                break

                        if is_splitter:
                            s["kind"] = "splitter"

        SCAN_STATE.update(
            {
                "status": "done",
                "shapes": deduped,
                "boundary": boundary_pts,
            }
        )

    except Exception as e:
        import traceback

        traceback.print_exc()
        SCAN_STATE.update({"status": "error", "error": str(e)})


@app.route("/api/scan_equipment", methods=["POST"])
def api_scan_equipment():
    data = request.get_json()
    dxf_path = data.get("dxf_path", "") or state.get("dxf_path", "")
    boundary_layer = data.get("boundary_layer") or None

    if not dxf_path:
        return jsonify({"error": "No DXF loaded"}), 400

    t = threading.Thread(
        target=_run_full_scan,
        args=(dxf_path, boundary_layer),
        daemon=True,
    )
    t.start()
    return jsonify({"ok": True})


@app.route("/api/scan_status")
def api_scan_status():
    return jsonify(
        {
            "status": SCAN_STATE["status"],
            "error": SCAN_STATE["error"],
            "progress": SCAN_STATE["progress"],
            "total": SCAN_STATE["total"],
            "count": len(SCAN_STATE["shapes"]),
        }
    )


@app.route("/api/scan_results")
def api_scan_results():
    segs = [
        {
            "x1": s.x1,
            "y1": s.y1,
            "x2": s.x2,
            "y2": s.y2,
            "is_hatch": getattr(s, "is_hatch", False),
        }
        for s in state["segments"]
    ]
    return jsonify(
        {
            "shapes": SCAN_STATE["shapes"],
            "boundary": SCAN_STATE["boundary"],
            "segments": segs,
        }
    )


# ─────────────────────────────────────────────────────────────────────────────
# POLE DETECTION
# ─────────────────────────────────────────────────────────────────────────────
import poleid as _poleid
from app_python.services.pole_ocr import ocr_pole

POLE_STATE: Dict = {
    "status": "idle",
    "error": None,
    "tags": [],
    "layer": None,
    "dxf_path": None,
    "progress": 0,
    "total": 0,
}

POLE_CONFIG = _poleid.PoleIdConfig(
    include_text=True,
    include_mtext=True,
    filter_text_by_regex=True,
    include_stroke=True,
    use_circle_markers=False,
    require_circle_match=False,
    max_dist_factor=4.0,
    default_text_height=0.25,
    stroke_connect_tol=0.20,
    stroke_min_total_length=0.30,
    stroke_min_segments=4,
    stroke_min_bbox_w=0.05,
    stroke_min_bbox_h=0.05,
    stroke_max_aspect=20.0,
    stroke_max_dom_dir=0.97,
    stroke_max_endpoints=24,
    stroke_placeholder_prefix="POLE",
)

OCR_WORKERS = 4


#: A circle further than this multiple of the median pole spacing from every
#: labelled pole carries no tag of its own.
UNTAGGED_CIRCLE_SPACING_RATIO = 0.3


def _untagged_pole_circles(doc, layer_names: list, tags: list) -> list:
    """Poles the drafter drew but never labelled.

    A pole symbol on the pole layer is a pole whether or not anyone typed a code
    beside it, and the cable ends there just the same. Skipping them let spans
    run straight through a pole and join the two on either side instead — the
    long span across half a node that the field reported.

    They are marked ``needs_review`` so the operator can name them; the
    derivation already treats them as span boundaries either way.
    """
    labelled = [(t["cx"], t["cy"]) for t in tags]
    centres = []
    for space in [doc.modelspace()] + [
        lay for lay in doc.layouts if lay.name.lower() != "model"
    ]:
        for e in space:
            if getattr(e.dxf, "layer", None) not in layer_names:
                continue
            if e.dxftype() != "CIRCLE":
                continue
            centres.append((float(e.dxf.center.x), float(e.dxf.center.y)))
    if not centres:
        return []

    # Spacing comes from the labelled poles, so an unlabelled cluster cannot
    # drag the threshold down and hide itself.
    spacing = _median_spacing(labelled) if len(labelled) > 1 else 0.0
    if spacing <= 0:
        spacing = _median_spacing(centres)
    if spacing <= 0:
        return []
    min_gap = spacing * UNTAGGED_CIRCLE_SPACING_RATIO

    next_id = max((t["pole_id"] for t in tags), default=-1) + 1
    extra = []
    for cx, cy in centres:
        if labelled and min(math.hypot(cx - x, cy - y) for x, y in labelled) <= min_gap:
            continue
        extra.append(
            {
                "pole_id": next_id,
                # NPT is the field crew's own term for a pole with no printed
                # tag, and it is what the AsBuilt side expects as an index.
                "name": f"NPT-{next_id:03d}",
                "cx": round(cx, 4),
                "cy": round(cy, 4),
                "bbox": [round(cx, 4), round(cy, 4), round(cx, 4), round(cy, 4)],
                "layer": layer_names[0] if layer_names else None,
                "source": "circle",
                "crop_b64": None,
                "ocr_conf": None,
                "needs_review": True,
                "untagged": True,
            }
        )
        labelled.append((cx, cy))
        next_id += 1

    if extra:
        print(f"[poles] {len(extra)} untagged pole circle(s) added from the pole layer")
    return extra


def _median_spacing(points: list) -> float:
    """Median nearest-neighbour distance for a set of points."""
    if len(points) < 2:
        return 0.0
    nn = []
    for i, a in enumerate(points):
        best = min(
            (math.hypot(a[0] - b[0], a[1] - b[1]) for j, b in enumerate(points) if i != j),
            default=float("inf"),
        )
        if best < float("inf"):
            nn.append(best)
    if not nn:
        return 0.0
    nn.sort()
    mid = len(nn) // 2
    return nn[mid] if len(nn) % 2 else (nn[mid - 1] + nn[mid]) / 2.0


def _run_pole_scan(dxf_path: str, layer_names: list[str]) -> None:
    try:
        combined = ", ".join(layer_names)
        POLE_STATE.update(
            {
                "status": "processing",
                "error": None,
                "tags": [],
                "layer": combined,
                "progress": 0,
                "total": 0,
            }
        )

        doc = ezdxf.readfile(dxf_path)
        placeholder_prefix = (POLE_CONFIG.stroke_placeholder_prefix or "POLE").upper()

        all_tags = []
        all_ocr_queue = []
        global_pole_id = 0
        tags_lock = threading.Lock()

        for layer_name in layer_names:
            matches = _poleid.find_pole_labels(doc, layer_name, config=POLE_CONFIG)
            layer_segs = extract_stroke_segments(doc, layer_name, include_circles=False)

            for lab, _circ in matches:
                bbox = list(lab.bbox) if lab.bbox else [lab.x, lab.y, lab.x, lab.y]
                source = getattr(lab, "source", "unknown")
                display_name = _poleid.clean_label(lab.text)
                is_placeholder = source == "stroke" and display_name.upper().startswith(placeholder_prefix)

                if is_placeholder:
                    all_ocr_queue.append((global_pole_id, lab, bbox, source, layer_name, layer_segs))
                else:
                    all_tags.append(
                        {
                            "pole_id": global_pole_id,
                            "name": display_name,
                            "cx": round(lab.x, 4),
                            "cy": round(lab.y, 4),
                            "bbox": [round(v, 4) for v in bbox],
                            "layer": layer_name,
                            "source": source,
                            "crop_b64": None,
                            "ocr_conf": None,
                            "needs_review": False,
                        }
                    )
                global_pole_id += 1

            with tags_lock:
                POLE_STATE["total"] = global_pole_id
                POLE_STATE["progress"] = len(all_tags)
                POLE_STATE["tags"] = list(all_tags)

        def _ocr_one(args):
            pole_id, lab, bbox, source, layer_name, layer_segs = args
            display_name = _poleid.clean_label(lab.text)
            crop_b64 = None
            ocr_conf = None
            needs_review = False

            try:
                # The label's own cluster when the detector provides it; the
                # whole layer only as a fallback. Rasterising the layer put the
                # pole circle and neighbouring IDs into every crop.
                label_segs = lab.segments if getattr(lab, "segments", None) else layer_segs
                result = ocr_pole(label_segs, tuple(bbox))
                if result.crop_png:
                    crop_b64 = base64.b64encode(result.crop_png).decode("ascii")
                ocr_conf = result.confidence

                if result.accepted and result.text:
                    display_name = result.text
                    needs_review = False
                else:
                    needs_review = True

            except Exception:
                needs_review = True

            return {
                "pole_id": pole_id,
                "name": display_name,
                "cx": round(lab.x, 4),
                "cy": round(lab.y, 4),
                "bbox": [round(v, 4) for v in bbox],
                "layer": layer_name,
                "source": source,
                "crop_b64": crop_b64,
                "ocr_conf": ocr_conf,
                "needs_review": needs_review,
            }

        if all_ocr_queue:
            with ThreadPoolExecutor(max_workers=OCR_WORKERS) as pool:
                futures = {pool.submit(_ocr_one, args): args for args in all_ocr_queue}
                for future in as_completed(futures):
                    try:
                        tag = future.result()
                    except Exception:
                        args = futures[future]
                        tag = {
                            "pole_id": args[0],
                            "name": _poleid.clean_label(args[1].text),
                            "cx": round(args[1].x, 4),
                            "cy": round(args[1].y, 4),
                            "bbox": [round(v, 4) for v in args[2]],
                            "layer": args[4],
                            "source": args[3],
                            "crop_b64": None,
                            "ocr_conf": None,
                            "needs_review": True,
                        }
                    with tags_lock:
                        all_tags.append(tag)
                        all_tags.sort(key=lambda t: t["pole_id"])
                        POLE_STATE["tags"] = list(all_tags)
                        POLE_STATE["progress"] = len(all_tags)

        all_tags.extend(_untagged_pole_circles(doc, layer_names, all_tags))
        all_tags.sort(key=lambda t: t["pole_id"])
        POLE_STATE.update(
            {
                "status": "done",
                "tags": all_tags,
                "progress": len(all_tags),
            }
        )

    except Exception as exc:
        import traceback

        traceback.print_exc()
        POLE_STATE.update({"status": "error", "error": str(exc)})


@app.route("/api/dxf_segments_no_circles")
def api_dxf_segments_no_circles():
    dxf_path = state.get("dxf_path")
    if not dxf_path:
        return jsonify({"error": "No DXF loaded"}), 400
    try:
        doc = ezdxf.readfile(dxf_path)
        layers = list_layers(dxf_path)
        all_segments = {}
        for layer in layers:
            segs = extract_stroke_segments(doc, layer, include_circles=False)
            all_segments[layer] = [
                {
                    "x1": s.x1,
                    "y1": s.y1,
                    "x2": s.x2,
                    "y2": s.y2,
                    "is_hatch": getattr(s, "is_hatch", False),
                }
                for s in segs
            ]
        return jsonify({"layers": layers, "segments": all_segments})
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/pole_tags")
def api_pole_tags():
    return jsonify(
        {
            "status": POLE_STATE["status"],
            "error": POLE_STATE["error"],
            "layer": POLE_STATE["layer"],
            "count": len(POLE_STATE["tags"]),
            "progress": POLE_STATE.get("progress", 0),
            "total": POLE_STATE.get("total", 0),
            "tags": POLE_STATE["tags"],
        }
    )


@app.route("/api/pole_tags/scan", methods=["POST"])
def api_pole_tags_scan():
    data = request.get_json()
    dxf_path = data.get("dxf_path", "") or state.get("dxf_path", "")
    layer_name = data.get("layer", "")
    layers = data.get("layers", [])
    if not layers and layer_name:
        layers = [layer_name]

    if not dxf_path:
        return jsonify({"error": "No DXF loaded"}), 400
    if not layers:
        return jsonify({"error": "at least one layer is required"}), 400

    t = threading.Thread(
        target=_run_pole_scan,
        args=(dxf_path, layers),
        daemon=True,
    )
    t.start()
    return jsonify({"ok": True, "layers": layers})


# ─────────────────────────────────────────────────────────────────────────────
# EQUIPMENT KIND → NAME MAPPING
# ─────────────────────────────────────────────────────────────────────────────


def kind_to_equipment_name(kind: str, layer: str = "") -> str:
    if kind == "circle":
        return "2-Way Tap"
    if kind == "splitter":
        return "Splitter"
    if kind == "square":
        return "4-Way Tap"
    if kind == "hexagon":
        return "8-Way Tap"
    if kind == "triangle":
        return "Line Extender"
    if kind == "rectangle":
        l = layer.lower()
        if "node" in l:
            return "Node"
        if "amp" in l or "amplifier" in l:
            return "Amplifier"
        return "Node/Amplifier"
    return kind.capitalize()


def _make_excel_styles():
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    return {
        "header_fill": PatternFill("solid", fgColor="1A3A5C"),
        "header_font": Font(bold=True, color="FFFFFF", name="Calibri"),
        "review_fill": PatternFill("solid", fgColor="FFF3CD"),
        "ok_fill": PatternFill("solid", fgColor="D4EDDA"),
        "sum_fill": PatternFill("solid", fgColor="E8EAF6"),
        "thin": Side(style="thin", color="CCCCCC"),
    }


def _make_border(styles):
    b = styles["thin"]
    from openpyxl.styles import Border

    return Border(left=b, right=b, top=b, bottom=b)


def _write_header_row(ws, headers, col_widths, styles):
    from openpyxl.styles import Alignment

    border = _make_border(styles)
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 22


def _write_footer_row(ws, row_num, cols, styles):
    from openpyxl.styles import Alignment, Font

    border = _make_border(styles)
    for col, val in cols:
        c = ws.cell(row=row_num, column=col, value=val)
        c.font = Font(bold=True)
        c.fill = styles["sum_fill"]
        c.border = border
        c.alignment = Alignment(horizontal="center")
    return c


def export_excel(results, dxf_path):
    try:
        import openpyxl
        from openpyxl.styles import Alignment
    except ImportError:
        return None, "openpyxl not installed."

    styles = _make_excel_styles()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Digit Results"
    border = _make_border(styles)
    dxf_name = Path(dxf_path).name
    headers = [
        "Digit ID",
        "Predicted Value",
        "Corrected Value",
        "Final Value",
        "Confidence %",
        "Needs Review",
        "Center X",
        "Center Y",
        "DXF File",
    ]
    col_widths = [10, 16, 16, 14, 14, 14, 12, 12, 30]
    _write_header_row(ws, headers, col_widths, styles)

    total_sum = 0
    for ri, r in enumerate(results, 2):
        final_val = (
            r["corrected_value"] if r["corrected_value"] is not None else r["value"]
        )
        try:
            total_sum += int(final_val)
        except Exception:
            pass
        row_data = [
            r["digit_id"],
            r["value"],
            r["corrected_value"] or "",
            final_val,
            round(r["confidence"] * 100, 1),
            "Yes" if r["needs_review"] else "No",
            round(r["center_x"], 4),
            round(r["center_y"], 4),
            dxf_name,
        ]
        fill = styles["review_fill"] if r["needs_review"] else styles["ok_fill"]
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

    sum_row = len(results) + 2
    ws.cell(row=sum_row, column=1, value="TOTAL").font = styles["header_font"]
    sum_cell = ws.cell(row=sum_row, column=4, value=total_sum)
    sum_cell.font = styles["header_font"]
    for ci in range(1, len(headers) + 1):
        c = ws.cell(row=sum_row, column=ci)
        c.fill = styles["sum_fill"]
        c.border = border

    ws2 = wb.create_sheet(title="Summary")
    _write_header_row(ws2, ["Digit ID", "Final Value"], [12, 16], styles)

    for ri, r in enumerate(results, 2):
        final_val = (
            r["corrected_value"] if r["corrected_value"] is not None else r["value"]
        )
        c1 = ws2.cell(row=ri, column=1, value=r["digit_id"])
        c2 = ws2.cell(row=ri, column=2, value=final_val)
        for c in (c1, c2):
            c.alignment = Alignment(horizontal="center")
            c.border = border

    tr = len(results) + 2
    _write_footer_row(ws2, tr, [(1, "TOTAL"), (2, total_sum)], styles)

    out_path = os.path.join(os.getcwd(), Path(dxf_path).stem + "_results.xlsx")
    wb.save(out_path)
    return out_path, None


def export_equipment_excel(shapes: list, dxf_path: str) -> tuple:
    try:
        import openpyxl
        from openpyxl.styles import Alignment
    except ImportError:
        return None, "openpyxl not installed."

    styles = _make_excel_styles()
    border = _make_border(styles)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Equipment"

    title_cell = ws.cell(row=1, column=1, value="Equipment Summary")
    title_cell.font = styles["header_font"]
    title_cell.fill = styles["header_fill"]
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 22

    kind_counts: Dict[str, int] = {}
    for sh in shapes:
        ek = kind_to_equipment_name(sh["kind"], sh.get("layer", ""))
        kind_counts[ek] = kind_counts.get(ek, 0) + 1

    ws.cell(row=2, column=1, value="Equipment").fill = styles["header_fill"]
    ws.cell(row=2, column=1, value="Equipment").font = styles["header_font"]
    ws.cell(row=2, column=1).border = border
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=2, column=2, value="Count").fill = styles["header_fill"]
    ws.cell(row=2, column=2, value="Count").font = styles["header_font"]
    ws.cell(row=2, column=2).border = border
    ws.cell(row=2, column=2).alignment = Alignment(horizontal="center")

    for ri, (ek, count) in enumerate(sorted(kind_counts.items()), 3):
        c1 = ws.cell(row=ri, column=1, value=ek)
        c2 = ws.cell(row=ri, column=2, value=count)
        for c in (c1, c2):
            c.border = border
            c.alignment = Alignment(horizontal="center")

    total_row = len(kind_counts) + 3
    _write_footer_row(ws, total_row, [(1, "TOTAL"), (2, len(shapes))], styles)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 10

    list_start = total_row + 2
    eq_headers = ["#", "Equipment", "Shape", "Layer", "Center X", "Center Y"]
    eq_widths = [8, 18, 12, 30, 12, 12]
    for ci, (h, w) in enumerate(zip(eq_headers, eq_widths), 1):
        cell = ws.cell(row=list_start, column=ci, value=h)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = w

    for ri, sh in enumerate(shapes, list_start + 1):
        row_data = [
            sh["shape_id"] + 1,
            kind_to_equipment_name(sh["kind"], sh.get("layer", "")),
            sh["kind"].capitalize(),
            sh.get("layer", ""),
            round(sh["cx"], 4),
            round(sh["cy"], 4),
        ]
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

    out_path = os.path.join(os.getcwd(), Path(dxf_path).stem + "_equipment.xlsx")
    wb.save(out_path)
    return out_path, None


def export_all_excel(
    results: list, shapes: list, poles: list, cable_spans: list, dxf_path: str
) -> tuple:
    try:
        import openpyxl
        from openpyxl.styles import Alignment
    except ImportError:
        return None, "openpyxl not installed."

    styles = _make_excel_styles()
    border = _make_border(styles)
    wb = openpyxl.Workbook()
    dxf_name = Path(dxf_path).stem

    # ── Sheet 1: OCR Results
    ws1 = wb.active
    ws1.title = "OCR Results"
    _write_header_row(
        ws1,
        [
            "Digit ID",
            "Predicted Value",
            "Corrected Value",
            "Final Value",
            "Confidence %",
            "Needs Review",
            "Center X",
            "Center Y",
        ],
        [10, 16, 16, 14, 14, 14, 12, 12],
        styles,
    )
    total_sum = 0
    for ri, r in enumerate(results, 2):
        final_val = (
            r["corrected_value"] if r["corrected_value"] is not None else r["value"]
        )
        try:
            total_sum += int(final_val)
        except:
            pass
        row_data = [
            r["digit_id"],
            r["value"],
            r["corrected_value"] or "",
            final_val,
            round(r["confidence"] * 100, 1),
            "Yes" if r["needs_review"] else "No",
            round(r["center_x"], 4),
            round(r["center_y"], 4),
        ]
        fill = styles["review_fill"] if r["needs_review"] else styles["ok_fill"]
        for ci, val in enumerate(row_data, 1):
            cell = ws1.cell(row=ri, column=ci, value=val)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
    sum_row = len(results) + 2
    _write_footer_row(
        ws1,
        sum_row,
        [(1, "TOTAL"), (4, total_sum), (8, dxf_name + "_results.xlsx")],
        styles,
    )

    # ── Sheet 2: Equipment
    if shapes:
        ws2 = wb.create_sheet(title="Equipment")
        title_cell = ws2.cell(row=1, column=1, value="Equipment Summary")
        title_cell.font = styles["header_font"]
        title_cell.fill = styles["header_fill"]
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws2.merge_cells("A1:B1")
        ws2.row_dimensions[1].height = 22

        kind_counts: Dict[str, int] = {}
        for sh in shapes:
            ek = kind_to_equipment_name(sh["kind"], sh.get("layer", ""))
            kind_counts[ek] = kind_counts.get(ek, 0) + 1

        ws2.cell(row=2, column=1, value="Equipment").fill = styles["header_fill"]
        ws2.cell(row=2, column=1, value="Equipment").font = styles["header_font"]
        ws2.cell(row=2, column=1).border = border
        ws2.cell(row=2, column=1).alignment = Alignment(horizontal="center")
        ws2.cell(row=2, column=2, value="Count").fill = styles["header_fill"]
        ws2.cell(row=2, column=2, value="Count").font = styles["header_font"]
        ws2.cell(row=2, column=2).border = border
        ws2.cell(row=2, column=2).alignment = Alignment(horizontal="center")

        for ri, (ek, count) in enumerate(sorted(kind_counts.items()), 3):
            c1 = ws2.cell(row=ri, column=1, value=ek)
            c2 = ws2.cell(row=ri, column=2, value=count)
            for c in (c1, c2):
                c.border = border
                c.alignment = Alignment(horizontal="center")

        eq_total_row = len(kind_counts) + 3
        _write_footer_row(ws2, eq_total_row, [(1, "TOTAL"), (2, len(shapes))], styles)
        ws2.column_dimensions["A"].width = 20
        ws2.column_dimensions["B"].width = 10

        list_start = eq_total_row + 2
        eq_headers = ["#", "Equipment", "Shape", "Layer", "Center X", "Center Y"]
        eq_widths = [8, 18, 12, 30, 12, 12]
        for ci, (h, w) in enumerate(zip(eq_headers, eq_widths), 1):
            cell = ws2.cell(row=list_start, column=ci, value=h)
            cell.fill = styles["header_fill"]
            cell.font = styles["header_font"]
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
            ws2.column_dimensions[cell.column_letter].width = w

        for ri, sh in enumerate(shapes, list_start + 1):
            row_data = [
                sh["shape_id"] + 1,
                kind_to_equipment_name(sh["kind"], sh.get("layer", "")),
                sh["kind"].capitalize(),
                sh.get("layer", ""),
                round(sh["cx"], 4),
                round(sh["cy"], 4),
            ]
            for ci, val in enumerate(row_data, 1):
                cell = ws2.cell(row=ri, column=ci, value=val)
                cell.border = border
                cell.alignment = Alignment(horizontal="center")

    # ── Sheet 3: Poles
    if poles:
        ws3 = wb.create_sheet(title="Poles")
        import openpyxl

        pole_fill = openpyxl.styles.PatternFill("solid", fgColor="FEF3C7")
        _write_header_row(
            ws3,
            [
                "#",
                "Pole Name",
                "Layer",
                "Source",
                "Confidence",
                "X",
                "Y",
                "Latitude",
                "Longitude",
            ],
            [8, 28, 20, 10, 12, 12, 12, 14, 14],
            styles,
        )
        ws3.freeze_panes = "A2"

        for ri, tag in enumerate(poles, 2):
            lat = tag.get("map_latitude")
            lon = tag.get("map_longitude")
            row_data = [
                ri - 1,
                tag.get("name", ""),
                tag.get("layer", ""),
                tag.get("source", ""),
                f"{round(tag.get('ocr_conf', 0) * 100, 1)}%"
                if tag.get("ocr_conf") is not None
                else "—",
                round(tag.get("cx", 0), 4),
                round(tag.get("cy", 0), 4),
                round(lat, 6) if lat is not None else "",
                round(lon, 6) if lon is not None else "",
            ]
            for ci, val in enumerate(row_data, 1):
                cell = ws3.cell(row=ri, column=ci, value=val)
                cell.fill = pole_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = border
        pole_total_row = len(poles) + 2
        _write_footer_row(ws3, pole_total_row, [(1, "TOTAL"), (2, len(poles))], styles)

    # ── Sheet 4: Cable Spans
    if cable_spans:
        ws4 = wb.create_sheet(title="Cable Spans")
        _write_header_row(
            ws4,
            [
                "Span #",
                "From Pole",
                "To Pole",
                "Layer",
                "Length",
                "Meter Value",
                "Center X",
                "Center Y",
            ],
            [10, 12, 12, 20, 12, 14, 12, 12],
            styles,
        )
        for ri, span in enumerate(cable_spans, 2):
            row_data = [
                span.get("span_id", ri - 2) + 1,
                span.get("from_pole", ""),
                span.get("to_pole", ""),
                span.get("layer", ""),
                round(span.get("total_length", 0), 4),
                span.get("meter_value") if span.get("meter_value") is not None else "—",
                round(span.get("cx", 0), 4),
                round(span.get("cy", 0), 4),
            ]
            for ci, val in enumerate(row_data, 1):
                cell = ws4.cell(row=ri, column=ci, value=val)
                cell.alignment = Alignment(horizontal="center")
                cell.border = border
        span_total_row = len(cable_spans) + 2
        _write_footer_row(
            ws4,
            span_total_row,
            [(1, "TOTAL"), (5, sum(s.get("total_length", 0) for s in cable_spans))],
            styles,
        )

    out_path = os.path.join(os.getcwd(), dxf_name + "_full_report.xlsx")
    wb.save(out_path)
    print(f"[export] Excel saved to {out_path}")
    return out_path, None


def export_poles_excel(tags: list, dxf_path: str) -> tuple:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError:
        return None, "openpyxl not installed."

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pole IDs"

    header_fill = PatternFill("solid", fgColor="7C4A00")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri")
    row_fill = PatternFill("solid", fgColor="FEF3C7")
    total_fill = PatternFill("solid", fgColor="E8EAF6")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["#", "Pole Name"]
    col_widths = [8, 28]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    for ri, tag in enumerate(tags, 2):
        row_data = [ri - 1, tag.get("name", "")]
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

    total_row = len(tags) + 2
    for col, val in [(1, "TOTAL"), (2, len(tags))]:
        c = ws.cell(row=total_row, column=col, value=val)
        c.font = Font(bold=True)
        c.fill = total_fill
        c.border = border
        c.alignment = Alignment(horizontal="center")

    out_path = os.path.join(os.getcwd(), Path(dxf_path).stem + "_pole_ids.xlsx")
    wb.save(out_path)
    return out_path, None


def _find_pole_layer_names(layers: List[str]) -> List[str]:
    patterns = ["pole", "poleid", "pole_id", "pole id", "tag", "label", "stp"]
    matched = []
    seen = set()
    for layer in layers:
        ll = layer.lower()
        if any(p in ll for p in patterns):
            if layer not in seen:
                matched.append(layer)
                seen.add(layer)
    return matched


@app.route("/api/pole_tags/auto_scan", methods=["POST"])
def api_pole_tags_auto_scan():
    data = request.get_json()
    dxf_path = data.get("dxf_path", "") or state.get("dxf_path", "")
    all_layers = data.get("layers", [])

    if not dxf_path:
        return jsonify({"error": "No DXF path provided"}), 400

    if (
        POLE_STATE.get("status") in ("processing", "done")
        and POLE_STATE.get("dxf_path") == dxf_path
    ):
        return jsonify({"ok": True, "skipped": True, "layers": POLE_STATE.get("layer")})

    layer_names = _find_pole_layer_names(all_layers)
    if not layer_names:
        return jsonify(
            {"ok": False, "reason": "No pole layer detected in this drawing"}
        )

    POLE_STATE["dxf_path"] = dxf_path

    t = threading.Thread(
        target=_run_pole_scan,
        args=(dxf_path, layer_names),
        daemon=True,
    )
    t.start()
    return jsonify({"ok": True, "layers": layer_names})


# ─────────────────────────────────────────────────────────────────────────────
# FLASK ROUTES
# ─────────────────────────────────────────────────────────────────────────────


@app.route("/")
def index():
    return send_from_directory(app.static_folder, "index.html")


@app.route("/api/status")
def api_status():
    return jsonify(
        {
            "status": state["status"],
            "progress": state["progress"],
            "total": state["total"],
            "error": state["error"],
            "step": state.get("step", 0),
            "step_label": state.get("step_label", ""),
        }
    )


@app.route("/api/results")
def api_results():
    segs = [
        {
            "x1": s.x1,
            "y1": s.y1,
            "x2": s.x2,
            "y2": s.y2,
            "is_hatch": getattr(s, "is_hatch", False),
        }
        for s in state["segments"]
    ]
    return jsonify({"results": state["results"], "segments": segs})


@app.route("/api/check_model")
def api_check_model():
    try:
        import easyocr

        return jsonify(
            {"ok": True, "engine": "easyocr", "cached": _easyocr_reader is not None}
        )
    except ImportError:
        return jsonify(
            {"ok": False, "engine": "easyocr", "error": "easyocr not installed"}
        )


# ─────────────────────────────────────────────────────────────────────────────
# FILE PLATFORM
# ─────────────────────────────────────────────────────────────────────────────

UPLOADS_DIR = Path("uploads")
INDEX_FILE = UPLOADS_DIR / "index.json"

# Track async PDF conversion jobs
# job_id -> { status: "pending"|"converting"|"done"|"error", path?, error? }
CONVERT_JOBS: Dict[str, dict] = {}

# ── Precompute job queue ──────────────────────────────────────────────────────
# checksum -> { status: "queued"|"processing"|"done"|"error", session_id?, error? }
_PRECOMPUTE_JOBS: Dict[str, dict] = {}
_PRECOMPUTE_QUEUE: "queue.Queue[str]" = None  # lazily initialised below


import queue as _queue_mod

_PRECOMPUTE_QUEUE = _queue_mod.Queue()


def _precompute_worker():
    """Serial background worker that processes one DXF precompute job at a time."""
    while True:
        checksum = _PRECOMPUTE_QUEUE.get()
        job = _PRECOMPUTE_JOBS.get(checksum)
        if not job:
            continue
        dxf_path = job.get("dxf_path", "")
        try:
            _run_precompute(checksum, dxf_path)
        except Exception as e:
            import traceback
            traceback.print_exc()
            _PRECOMPUTE_JOBS[checksum].update({"status": "error", "error": str(e)})
        finally:
            _PRECOMPUTE_QUEUE.task_done()


threading.Thread(target=_precompute_worker, daemon=True).start()


def _run_precompute(checksum: str, dxf_path: str):
    """
    Full headless pipeline for one DXF file:
      strand OCR → poles → equipment → cable spans → dxf segments
    Writes everything to Supabase via session_store.
    """
    job = _PRECOMPUTE_JOBS[checksum]
    job.update({"status": "processing"})

    # ── get/create Supabase project+session ───────────────────────────────────
    project_id, session_id, is_new = get_or_create_project_session(dxf_path, checksum)

    if session_id and not is_new and session_has_user_edits(session_id):
        print(f"[precompute] {dxf_path}: session {session_id} has user edits — skipping overwrite")
        job.update({"status": "done", "session_id": session_id, "skipped": True})
        return

    layers = list_layers(dxf_path)
    doc = ezdxf.readfile(dxf_path)

    # ── dxf segments (all layers) ─────────────────────────────────────────────
    all_segments_by_layer: Dict[str, list] = {}
    for lyr in layers:
        segs = extract_stroke_segments(doc, lyr, include_circles=False)
        all_segments_by_layer[lyr] = [
            {"x1": s.x1, "y1": s.y1, "x2": s.x2, "y2": s.y2,
             "is_hatch": getattr(s, "is_hatch", False)}
            for s in segs
        ]

    # ── strand OCR (auto-detect strand layer) ─────────────────────────────────
    strand_patterns = ["strand", "wire", "drop", "fiber", "fibre"]
    strand_layers = [
        l for l in layers
        if any(p in l.lower() for p in strand_patterns)
    ]
    # Fallback: pick a layer that has the most digit-like clusters
    if not strand_layers:
        strand_layers = [layers[0]] if layers else []

    digit_results: list = []
    if strand_layers:
        all_segs: list = []
        for lyr in strand_layers:
            segs_raw = extract_stroke_segments(doc, lyr, include_circles=False)
            all_segs.extend(segs_raw)

        if all_segs:
            scale = estimate_scale(all_segs)
            cable = cable_segment_indices(all_segs, scale=scale)
            clusters = cluster_segments(all_segs, tol=CONNECT_TOL * scale, ignore=cable)
            infos = analyze_clusters(all_segs, clusters, scale=scale)
            candidates = build_candidates_robust(all_segs, infos, scale=scale)

            if candidates:
                crops = [render_crop(all_segs, cand) for cand in candidates]
                from app_python.services.strand_recognizer import recognize_batch
                batch = recognize_batch(crops)
                for cand, crop, (value, conf) in zip(candidates, crops, batch):
                    cx = (cand.bbox[0] + cand.bbox[2]) / 2
                    cy = (cand.bbox[1] + cand.bbox[3]) / 2
                    needs_review = (not value) or (conf < _MIN_CONF)
                    digit_results.append({
                        "digit_id": cand.digit_id,
                        "value": value if value else "?",
                        "corrected_value": None,
                        "confidence": round(conf, 4),
                        "needs_review": needs_review,
                        "bbox": list(cand.bbox),
                        "center_x": cx,
                        "center_y": cy,
                    })
                digit_results = _post_ocr_validate(digit_results)

    # ── poles ─────────────────────────────────────────────────────────────────
    pole_layer_names = _find_pole_layer_names(layers)
    poles: list = []
    if pole_layer_names:
        global_pole_id = 0
        for layer_name in pole_layer_names:
            matches = _poleid.find_pole_labels(doc, layer_name, config=POLE_CONFIG)
            placeholder_prefix = (
                POLE_CONFIG.stroke_placeholder_prefix or "POLE"
            ).upper()
            for lab, _circ in matches:
                bbox = list(lab.bbox) if lab.bbox else [lab.x, lab.y, lab.x, lab.y]
                source = getattr(lab, "source", "text")
                display_name = _poleid.clean_label(lab.text)
                # This headless path never runs pole OCR, so a stroke label
                # still carrying the POLE_xxx placeholder has NOT been read.
                # It used to be stored with needs_review=False, and restored
                # sessions then presented hundreds of unread poles as
                # confirmed names.
                is_placeholder = (
                    source == "stroke"
                    and display_name.upper().startswith(placeholder_prefix)
                )
                poles.append({
                    "pole_id": global_pole_id,
                    "name": display_name,
                    "cx": round(lab.x, 4),
                    "cy": round(lab.y, 4),
                    "bbox": [round(v, 4) for v in bbox],
                    "layer": layer_name,
                    "source": source,
                    "ocr_conf": None,
                    "needs_review": is_placeholder,
                })
                global_pole_id += 1

    # ── equipment shapes ──────────────────────────────────────────────────────
    # Reuse the same detection logic as _run_full_scan but without touching SCAN_STATE
    KIND_LAYER_MAP = {
        "circle": ["splitter", "tapoff", "tap-off", "tap_off", "splt"],
        "hexagon": ["tapoff", "tap-off", "tap_off"],
        "rectangle": ["node", "amplifier", "amp"],
        "square": ["tapoff", "tap-off", "tap_off"],
        "triangle": ["extender", "extend"],
    }
    layer_kind_targets: Dict[str, list] = {}
    for layer in layers:
        l_lower = layer.lower()
        kinds_for_layer = [
            kind for kind, keywords in KIND_LAYER_MAP.items()
            if any(kw in l_lower for kw in keywords)
        ]
        if kinds_for_layer:
            layer_kind_targets[layer] = kinds_for_layer

    all_shapes: list = []
    for layer, allowed_kinds in layer_kind_targets.items():
        try:
            shapes = extract_equipment_shapes(doc, layer, **SHAPE_CONFIG)
            for s in shapes:
                if s.kind not in allowed_kinds:
                    continue
                all_shapes.append({
                    "shape_id": -1,
                    "kind": s.kind,
                    "bbox": list(s.bbox),
                    "cx": s.cx,
                    "cy": s.cy,
                    "layer": layer,
                })
        except Exception:
            pass

    # dedup + re-id
    DEDUP_EPS = 0.5
    all_shapes.sort(key=lambda s: (s["kind"], s["cx"], s["cy"]))
    deduped: list = []
    for s in all_shapes:
        if not deduped:
            deduped.append(s)
            continue
        prev = deduped[-1]
        if (s["kind"] == prev["kind"]
                and abs(s["cx"] - prev["cx"]) < DEDUP_EPS
                and abs(s["cy"] - prev["cy"]) < DEDUP_EPS):
            continue
        deduped.append(s)
    deduped.sort(key=lambda s: (-s["cy"], s["cx"]))
    for i, s in enumerate(deduped):
        s["shape_id"] = i

    # ── cable spans ───────────────────────────────────────────────────────────
    cable_layer_names = find_cable_layer_names(layers)
    cable_spans: list = []
    if cable_layer_names:
        cable_spans = build_cable_spans(doc, cable_layer_names, connect_tol=CABLE_CONNECT_TOL)
        if digit_results:
            cable_spans = assign_meter_values_to_spans(cable_spans, digit_results)

    # ── persist ───────────────────────────────────────────────────────────────
    if session_id:
        save_full_results(
            session_id,
            digit_results=digit_results,
            poles=poles,
            equipment_shapes=deduped,
            cable_spans=cable_spans,
            dxf_segments_by_layer=all_segments_by_layer,
            strand_layers=strand_layers,
            pole_layers=pole_layer_names,
            equipment_layers=list(layer_kind_targets.keys()),
        )

    job.update({
        "status": "done",
        "session_id": session_id,
        "counts": {
            "digits": len(digit_results),
            "poles": len(poles),
            "shapes": len(deduped),
            "spans": len(cable_spans),
        },
    })
    print(f"[precompute] done: {dxf_path}")


def _run_pdf_conversion(job_id: str, pdf_path: str, folder: str):
    """
    Background thread function for PDF-to-DXF conversion.
    Updates CONVERT_JOBS with progress and result.
    """
    try:
        CONVERT_JOBS[job_id]["status"] = "converting"
        print(f"[convert] Starting PDF conversion for job {job_id}: {pdf_path}")

        # Run AutoCAD conversion
        dxf_path_str = pdf_to_dxf_autocad(pdf_path)

        # Update file index
        fname = Path(dxf_path_str).name
        p = Path(dxf_path_str)
        data = _read_index()
        # Remove any existing entry with same path
        data["files"] = [
            f for f in data["files"] if _index_key(f["path"]) != _index_key(dxf_path_str)
        ]
        data["files"].append(
            {
                "name": fname,
                "path": dxf_path_str,
                "size": p.stat().st_size,
                "modified": int(p.stat().st_mtime),
                "folder": folder,
            }
        )
        _write_index(data)

        # Delete original PDF
        Path(pdf_path).unlink(missing_ok=True)

        # Mark job as done
        CONVERT_JOBS[job_id] = {
            "status": "done",
            "path": dxf_path_str,
            "name": fname,
        }
        print(f"[convert] Job {job_id} completed: {dxf_path_str}")

    except Exception as e:
        import traceback

        traceback.print_exc()
        CONVERT_JOBS[job_id] = {
            "status": "error",
            "error": str(e),
        }
        print(f"[convert] Job {job_id} failed: {e}")


def _index_key(path: str) -> str:
    """One file, one key.

    Paths reach the index from several places — direct upload, batch upload, PDF
    conversion — and arrive in different shapes: backslashes or forward slashes,
    relative to the working directory or fully qualified. Comparing the raw
    strings let one drawing be recorded under two identities, which surfaced as
    duplicate entries in the file list.
    """
    text = str(path or "").strip()
    if not text:
        return ""
    try:
        resolved = str(Path(text).resolve())
    except (OSError, ValueError):
        resolved = text
    return resolved.replace("\\", "/").rstrip("/").casefold()


def _dedupe_files(files: list) -> list:
    """Last entry wins, so a re-upload replaces rather than accumulates."""
    by_key: Dict[str, dict] = {}
    for f in files:
        by_key[_index_key(f.get("path", ""))] = f
    return list(by_key.values())


def _read_index() -> dict:
    UPLOADS_DIR.mkdir(exist_ok=True)
    if not INDEX_FILE.exists():
        return {"folders": [], "files": []}
    try:
        data = json.loads(INDEX_FILE.read_text(encoding="utf-8"))
    except Exception:
        return {"folders": [], "files": []}
    data["files"] = _dedupe_files(data.get("files", []))
    data.setdefault("folders", [])
    return data


def _write_index(data: dict) -> None:
    UPLOADS_DIR.mkdir(exist_ok=True)
    data["files"] = _dedupe_files(data.get("files", []))
    INDEX_FILE.write_text(json.dumps(data, indent=2), encoding="utf-8")


def _sync_index_sizes() -> dict:
    data = _read_index()
    kept = []
    for f in data["files"]:
        p = Path(f["path"])
        if p.exists():
            f["size"] = p.stat().st_size
            kept.append(f)
    data["files"] = kept
    _write_index(data)
    return data


@app.route("/api/files/list", methods=["GET"])
def api_files_list():
    data = _sync_index_sizes()
    folder_counts: Dict[str, int] = {f: 0 for f in data["folders"]}
    for file in data["files"]:
        folder = file.get("folder", "")
        if folder and folder in folder_counts:
            folder_counts[folder] += 1
    folders_out = [
        {"name": name, "fileCount": cnt} for name, cnt in folder_counts.items()
    ]
    return jsonify({"folders": folders_out, "files": data["files"]})


@app.route("/api/files/mkdir", methods=["POST"])
def api_files_mkdir():
    body = request.get_json() or {}
    name = (body.get("name") or "").strip()
    if not name:
        return jsonify({"error": "name is required"}), 400
    data = _read_index()
    if name not in data["folders"]:
        data["folders"].append(name)
        (UPLOADS_DIR / name).mkdir(parents=True, exist_ok=True)
        _write_index(data)
    return jsonify({"ok": True})


@app.route("/api/files/delete", methods=["POST"])
def api_files_delete():
    body = request.get_json() or {}
    path = (body.get("path") or "").strip()
    if not path:
        return jsonify({"error": "path is required"}), 400
    try:
        Path(path).resolve().relative_to(UPLOADS_DIR.resolve())
    except ValueError:
        return jsonify({"error": "Invalid path"}), 400
    Path(path).unlink(missing_ok=True)
    data = _read_index()
    data["files"] = [
        f for f in data["files"] if _index_key(f["path"]) != _index_key(path)
    ]
    _write_index(data)
    return jsonify({"ok": True})


@app.route("/api/files/rename", methods=["POST"])
def api_files_rename():
    body = request.get_json() or {}
    old_path_str = (body.get("path") or "").strip()
    new_name = (body.get("new_name") or "").strip()
    if not old_path_str or not new_name:
        return jsonify({"error": "path and new_name are required"}), 400
    old_path = Path(old_path_str)
    if not old_path.exists():
        return jsonify({"error": "File not found"}), 404
    suffix = old_path.suffix or ".dxf"
    if not new_name.lower().endswith(suffix.lower()):
        new_name = new_name + suffix
    new_path = old_path.parent / new_name
    old_path.rename(new_path)
    data = _read_index()
    for f in data["files"]:
        if _index_key(f["path"]) == _index_key(old_path_str):
            f["path"] = str(new_path)
            f["name"] = new_name
            break
    _write_index(data)
    return jsonify({"ok": True, "path": str(new_path)})


@app.route("/api/upload", methods=["POST"])
@app.route("/api/files/upload", methods=["POST"])
def api_upload():
    try:
        file = request.files.get("file")
        if not file:
            return jsonify({"error": "No file provided"}), 400

        fname = Path(file.filename).name if file.filename else "uploaded.dxf"
        folder = (request.form.get("folder") or "").strip()

        UPLOADS_DIR.mkdir(exist_ok=True)

        if folder:
            dest_dir = UPLOADS_DIR / folder
            dest_dir.mkdir(parents=True, exist_ok=True)
        else:
            dest_dir = UPLOADS_DIR

        save_path = str(dest_dir / fname)
        file.save(save_path)

        # Handle PDF files asynchronously to avoid timeout
        if save_path.lower().endswith(".pdf"):
            job_id = str(uuid.uuid4())
            CONVERT_JOBS[job_id] = {"status": "pending"}

            # Start conversion in background thread
            thread = threading.Thread(
                target=_run_pdf_conversion,
                args=(job_id, save_path, folder),
                daemon=True,
            )
            thread.start()

            print(f"[upload] PDF conversion job started: {job_id}")
            return jsonify(
                {
                    "converting": True,
                    "job_id": job_id,
                    "message": "PDF conversion started. Poll /api/files/convert-status for progress.",
                }
            )

        # DXF files are handled immediately
        p = Path(save_path)
        data = _read_index()
        data["files"] = [
            f for f in data["files"] if _index_key(f["path"]) != _index_key(save_path)
        ]
        data["files"].append(
            {
                "name": fname,
                "path": save_path,
                "size": p.stat().st_size,
                "modified": int(p.stat().st_mtime),
                "folder": folder,
            }
        )
        _write_index(data)

        # Enqueue headless precompute (Supabase must be configured)
        try:
            checksum = compute_checksum(save_path)
            existing = _PRECOMPUTE_JOBS.get(checksum, {})
            if existing.get("status") not in ("queued", "processing", "done"):
                _PRECOMPUTE_JOBS[checksum] = {"status": "queued", "dxf_path": save_path}
                _PRECOMPUTE_QUEUE.put(checksum)
                print(f"[upload] precompute queued for {fname} ({checksum[:8]})")
        except Exception as _pc_err:
            print(f"[upload] precompute enqueue failed (non-fatal): {_pc_err}")

        return jsonify({"path": save_path})

    except Exception as e:
        import traceback

        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/api/files/convert-status", methods=["GET"])
def api_convert_status():
    job_id = request.args.get("job_id")
    if not job_id:
        return jsonify({"error": "job_id is required"}), 400

    if job_id not in CONVERT_JOBS:
        return jsonify({"error": "Invalid or expired job_id"}), 404

    job = CONVERT_JOBS[job_id]

    return jsonify(job)


@app.route("/api/precompute/status", methods=["GET"])
def api_precompute_status():
    """Poll the precompute status for a DXF file (by path or checksum)."""
    dxf_path = request.args.get("dxf_path", "")
    checksum = request.args.get("checksum", "")

    if dxf_path and not checksum:
        try:
            checksum = compute_checksum(dxf_path)
        except Exception as e:
            return jsonify({"error": str(e)}), 400

    if not checksum:
        return jsonify({"error": "dxf_path or checksum is required"}), 400

    job = _PRECOMPUTE_JOBS.get(checksum)
    if not job:
        return jsonify({"status": "not_queued", "checksum": checksum})

    return jsonify({**job, "checksum": checksum})


@app.route("/api/precompute/trigger", methods=["POST"])
def api_precompute_trigger():
    """Manually trigger (or re-trigger) precompute for a DXF file."""
    data = request.get_json() or {}
    dxf_path = data.get("dxf_path", "")
    force = bool(data.get("force", False))

    if not dxf_path or not Path(dxf_path).exists():
        return jsonify({"error": "valid dxf_path is required"}), 400

    try:
        checksum = compute_checksum(dxf_path)
    except Exception as e:
        return jsonify({"error": str(e)}), 400

    existing = _PRECOMPUTE_JOBS.get(checksum, {})
    if existing.get("status") in ("queued", "processing") and not force:
        return jsonify({"ok": True, "queued": False, "status": existing["status"],
                        "checksum": checksum})

    _PRECOMPUTE_JOBS[checksum] = {"status": "queued", "dxf_path": dxf_path}
    _PRECOMPUTE_QUEUE.put(checksum)
    return jsonify({"ok": True, "queued": True, "checksum": checksum})


# ─────────────────────────────────────────────────────────────────────────────
# CHATBOT SUPPORT ENDPOINTS
# ─────────────────────────────────────────────────────────────────────────────


@app.route("/api/chat/summary", methods=["GET"])
def api_chat_summary():
    # Files
    index = _read_index()
    file_count = len(index.get("files", []))
    folders = index.get("folders", [])

    # Count files by folder
    folder_counts = {"root": 0}
    for f in folders:
        folder_counts[f] = 0
    for f in index.get("files", []):
        key = f.get("folder") or "root"
        folder_counts[key] = folder_counts.get(key, 0) + 1

    # Current file
    current_file = state.get("dxf_path")

    # OCR Results
    results = state.get("results", [])
    ocr_count = len(results)
    ocr_needing_review = len([r for r in results if r.get("needs_review")])

    total_strand_meters = 0
    for r in results:
        val = r.get("corrected_value") or r.get("value") or "0"
        if str(val).isdigit():
            total_strand_meters += int(val)

    # Poles
    poles = POLE_STATE.get("tags", [])
    pole_count = len(poles)
    poles_needing_review = len([p for p in poles if p.get("needs_review")])

    # Equipment
    shapes = SCAN_STATE.get("shapes", [])
    equipment_count = len(shapes)
    equipment_by_kind = {}
    for s in shapes:
        kind = s.get("kind", "unknown")
        equipment_by_kind[kind] = equipment_by_kind.get(kind, 0) + 1

    # Status
    ocr_status = state.get("status", "idle")
    pole_status = POLE_STATE.get("status", "idle")
    equipment_status = SCAN_STATE.get("status", "idle")

    return jsonify(
        {
            "ok": True,
            "data": {
                "files": {
                    "total": file_count,
                    "folders": folders,
                    "by_folder": folder_counts,
                },
                "current_file": current_file,
                "ocr": {
                    "count": ocr_count,
                    "needs_review": ocr_needing_review,
                    "total_strand_meters": total_strand_meters,
                    "status": ocr_status,
                },
                "poles": {
                    "count": pole_count,
                    "needs_review": poles_needing_review,
                    "status": pole_status,
                },
                "equipment": {
                    "count": equipment_count,
                    "by_kind": equipment_by_kind,
                    "status": equipment_status,
                },
            },
        }
    )


@app.route("/api/files/upload-batch", methods=["POST"])
def api_upload_batch():
    files = request.files.getlist("files")
    folder = (request.form.get("folder") or "").strip()

    if not files:
        return jsonify({"ok": False, "error": "No files provided"}), 400

    UPLOADS_DIR.mkdir(exist_ok=True)

    if folder:
        dest_dir = UPLOADS_DIR / folder
        dest_dir.mkdir(parents=True, exist_ok=True)
        # Ensure folder is in index
        data = _read_index()
        if folder not in data["folders"]:
            data["folders"].append(folder)
            _write_index(data)
    else:
        dest_dir = UPLOADS_DIR

    results = []
    for file in files:
        try:
            if not file.filename:
                results.append(
                    {"name": "unknown", "status": "error", "error": "No filename"}
                )
                continue

            fname = Path(file.filename).name
            ext = Path(fname).suffix.lower()

            if ext not in [".dxf", ".pdf"]:
                results.append(
                    {
                        "name": fname,
                        "status": "error",
                        "error": f"Unsupported file type: {ext}. Only .dxf and .pdf are supported.",
                    }
                )
                continue

            save_path = str(dest_dir / fname)
            file.save(save_path)

            # Handle PDF conversion asynchronously
            if ext == ".pdf":
                job_id = str(uuid.uuid4())
                CONVERT_JOBS[job_id] = {"status": "pending"}

                thread = threading.Thread(
                    target=_run_pdf_conversion,
                    args=(job_id, save_path, folder),
                    daemon=True,
                )
                thread.start()

                results.append(
                    {
                        "name": fname,
                        "status": "converting",
                        "job_id": job_id,
                    }
                )
            else:
                # DXF files - update index immediately
                p = Path(save_path)
                data = _read_index()
                data["files"] = [
                    f
                    for f in data["files"]
                    if _index_key(f["path"]) != _index_key(save_path)
                ]
                data["files"].append(
                    {
                        "name": fname,
                        "path": save_path,
                        "size": p.stat().st_size,
                        "modified": int(p.stat().st_mtime),
                        "folder": folder,
                    }
                )
                _write_index(data)

                results.append(
                    {
                        "name": fname,
                        "status": "success",
                        "path": save_path,
                    }
                )

        except Exception as e:
            import traceback

            traceback.print_exc()
            results.append(
                {
                    "name": file.filename or "unknown",
                    "status": "error",
                    "error": str(e),
                }
            )

    successful = len([r for r in results if r["status"] in ["success", "converting"]])

    return jsonify(
        {
            "ok": True,
            "results": results,
            "total": len(files),
            "successful": successful,
        }
    )


@app.route("/api/layers", methods=["POST"])
def api_layers():
    data = request.get_json()
    new_path = data.get("dxf_path", "")

    # Clear stale states if switching to new file
    if new_path and new_path != state.get("dxf_path"):
        _clear_all_states()
        state["dxf_path"] = new_path

    try:
        layers = list_layers(new_path)
        return jsonify({"layers": layers})
    except Exception as e:
        return jsonify({"error": str(e)}), 400


# <-- UPDATED TO ACCEPT 'LAYERS' LIST -->
@app.route("/api/run", methods=["POST"])
def api_run():
    data = request.get_json()
    state["dxf_path"] = data["dxf_path"]

    layers = data.get("layers", [])
    if not layers and "layer" in data:
        layers = [data["layer"]]

    state["layers"] = layers
    state["model_path"] = data.get("model_path", "")
    t = threading.Thread(
        target=run_pipeline,
        args=(data["dxf_path"], state["layers"], state["model_path"]),
        daemon=True,
    )
    t.start()
    return jsonify({"ok": True})


@app.route("/api/export", methods=["POST"])
def api_export():
    data = request.get_json() or {}
    corrections = data.get("corrections", {})
    for r in state["results"]:
        did = str(r["digit_id"])
        if did in corrections and corrections[did] is not None:
            r["corrected_value"] = corrections[did]
    path, err = export_excel(state["results"], state["dxf_path"])
    if err:
        return jsonify({"error": err}), 500
    return jsonify({"path": path})


@app.route("/api/export/equipment", methods=["POST"])
def api_export_equipment():
    shapes = SCAN_STATE.get("shapes", [])
    if not shapes:
        return jsonify({"error": "No equipment found. Run a scan first."}), 400
    path, err = export_equipment_excel(
        shapes, state.get("dxf_path") or "equipment_export"
    )
    if err:
        return jsonify({"error": err}), 500
    return jsonify({"path": path})


@app.route("/api/export/all", methods=["POST"])
def api_export_all():
    data = request.get_json() or {}
    corrections = data.get("corrections", {})
    for r in state["results"]:
        did = str(r["digit_id"])
        if did in corrections and corrections[did] is not None:
            r["corrected_value"] = corrections[did]
    shapes = SCAN_STATE.get("shapes", [])
    poles = POLE_STATE.get("tags", [])
    cable_spans = data.get("cable_spans", [])
    path, err = export_all_excel(
        state["results"],
        shapes,
        poles,
        cable_spans,
        state.get("dxf_path") or "full_report",
    )
    if err:
        return jsonify({"error": err}), 500

    return jsonify({"path": path})


@app.route("/api/export/polemaster", methods=["POST"])
def api_export_polemaster():
    """Push data to TelcoVantage Planner API (Pole Master)."""
    data = request.get_json() or {}
    corrections = data.get("corrections", {})

    # Apply corrections to OCR results
    for r in state["results"]:
        did = str(r["digit_id"])
        if did in corrections and corrections[did] is not None:
            r["corrected_value"] = corrections[did]

    shapes = SCAN_STATE.get("shapes", [])
    poles = POLE_STATE.get("tags", [])
    cable_spans = data.get("cable_spans", [])
    project_id = data.get("project_id")

    print(
        f"[polemaster] ENABLE_PLANNER_INTEGRATION: {ENABLE_PLANNER_INTEGRATION}, project_id: {project_id}"
    )
    print(f"[polemaster] Received {len(cable_spans)} cable spans, {len(poles)} poles")

    if cable_spans:
        sample = cable_spans[0]
        print(
            f"[polemaster] Sample span: from_pole={sample.get('from_pole')}, to_pole={sample.get('to_pole')}, "
            f"from_pole_id={sample.get('from_pole_id')}, to_pole_id={sample.get('to_pole_id')}"
        )
        # Log all unique from_pole_id / to_pole_id values to diagnose spanning issues
        seen_from = set()
        seen_to = set()
        for s in cable_spans:
            fid = s.get("from_pole_id")
            tid = s.get("to_pole_id")
            if fid is not None: seen_from.add(fid)
            if tid is not None: seen_to.add(tid)
        print(
            f"[polemaster] Unique from_pole_ids: {sorted(map(str, seen_from))}"
        )
        print(
            f"[polemaster] Unique to_pole_ids: {sorted(map(str, seen_to))}"
        )
        null_from = sum(1 for s in cable_spans if s.get("from_pole_id") is None)
        null_to = sum(1 for s in cable_spans if s.get("to_pole_id") is None)
        if null_from or null_to:
            print(f"[polemaster] WARNING: {null_from} spans have null from_pole_id, {null_to} have null to_pole_id")

    if not ENABLE_PLANNER_INTEGRATION:
        return jsonify({"error": "Planner integration is disabled"}), 400

    if project_id is None:
        project_id = DEFAULT_PROJECT_ID
        print(f"[polemaster] Using default project_id: {project_id}")

    if project_id is None:
        return jsonify(
            {"error": "No project_id provided and no default configured"}
        ), 400

    try:
        push_result = push_to_planner(
            state.get("dxf_path"),
            poles,
            cable_spans,
            shapes,
            state.get("results", []),
            int(project_id),
        )
        return jsonify({"success": True, "result": push_result})
    except Exception as e:
        import traceback

        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/api/download")
def api_download():
    fpath = request.args.get("file", "")
    if not fpath or not Path(fpath).exists():
        return "File not found", 404
    return send_file(
        fpath,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=Path(fpath).name,
    )


@app.route("/api/dxf_segments")
def api_dxf_segments():
    # Support optional dxg_path param for file switching
    new_path = request.args.get("dxf_path") or state.get("dxf_path")
    # Clear stale states if switching to new file
    if new_path and new_path != state.get("dxf_path"):
        _clear_all_states()
        state["dxf_path"] = new_path

    dxf_path = new_path
    hide_circles = request.args.get("hide_circles") == "1"
    if not dxf_path:
        return jsonify({"error": "No DXF loaded"}), 400
    try:
        # Pure function of the file — 22s of tessellation and 7 MB of JSON
        # were being rebuilt on every viewer mount.
        try:
            mtime = Path(dxf_path).stat().st_mtime
        except OSError:
            mtime = 0.0
        key = (dxf_path, hide_circles)
        cached = _DXF_SEGMENTS_CACHE.get(key)
        if cached and cached[0] == mtime:
            return app.response_class(cached[1], mimetype="application/json")

        doc = ezdxf.readfile(dxf_path)
        layers = list_layers(dxf_path)
        all_segments = {}
        for layer in layers:
            segs = extract_stroke_segments(doc, layer, include_circles=not hide_circles)
            all_segments[layer] = [
                {
                    "x1": s.x1,
                    "y1": s.y1,
                    "x2": s.x2,
                    "y2": s.y2,
                    "is_hatch": getattr(s, "is_hatch", False),
                }
                for s in segs
            ]
        body = json.dumps({"layers": layers, "segments": all_segments})
        _DXF_SEGMENTS_CACHE[key] = (mtime, body)
        return app.response_class(body, mimetype="application/json")
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/cable_spans")
def api_cable_spans():
    # Support optional dxf_path param for file switching
    new_path = request.args.get("dxf_path") or state.get("dxf_path")
    # Clear stale states if switching to new file
    if new_path and new_path != state.get("dxf_path"):
        _clear_all_states()
        state["dxf_path"] = new_path

    dxf_path = new_path
    if not dxf_path:
        return jsonify({"error": "No DXF loaded"}), 400
    try:
        # ?whole=true: the operator has not run the pole step in this session,
        # so serve the strand undivided regardless of what any earlier scan
        # left in POLE_STATE.
        if request.args.get("whole", "").lower() == "true":
            spans, cable_layers = _whole_cable_spans(dxf_path)
            return jsonify(
                {
                    "cable_layers": cable_layers,
                    "count": len(spans),
                    "spans": spans,
                    "poles": [],
                    "warnings": [],
                    "errors": [],
                    "status": "awaiting_poles",
                }
            )
        return jsonify(_span_response(dxf_path))
    except Exception as e:
        import traceback

        traceback.print_exc()
        return jsonify({"error": str(e)}), 400


@app.route("/api/planner/projects")
def api_planner_projects():
    """Fetch list of projects from Planner API."""
    try:
        from app_python.services.planner_auth import get_projects

        projects = get_projects()
        return jsonify({"projects": projects})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─────────────────────────────────────────────────────────────────────────────
# GEOREFERENCING  —  Map CAD poles to GPS coordinates
# ─────────────────────────────────────────────────────────────────────────────


def transform_coordinate(point, cad_p1, cad_p2, map_p1, map_p2):
    """Similarity transform: maps a CAD point (x,y) to GPS (lat,lon)
    using two anchor pairs."""
    dx_cad = cad_p2[0] - cad_p1[0]
    dy_cad = cad_p2[1] - cad_p1[1]
    dx_map = map_p2[1] - map_p1[1]  # Lon delta
    dy_map = map_p2[0] - map_p1[0]  # Lat delta

    dist_cad = math.hypot(dx_cad, dy_cad)
    dist_map = math.hypot(dx_map, dy_map)
    scale = dist_map / dist_cad if dist_cad > 0 else 1.0

    angle_cad = math.atan2(dy_cad, dx_cad)
    angle_map = math.atan2(dy_map, dx_map)
    angle_diff = angle_map - angle_cad

    vx = point[0] - cad_p1[0]
    vy = point[1] - cad_p1[1]

    new_x = (vx * math.cos(angle_diff) - vy * math.sin(angle_diff)) * scale
    new_y = (vx * math.sin(angle_diff) + vy * math.cos(angle_diff)) * scale

    final_lon = map_p1[1] + new_x
    final_lat = map_p1[0] + new_y
    return (final_lat, final_lon)


@app.route("/api/geocode")
def api_geocode():
    loc = request.args.get("loc", "")
    if not loc:
        return jsonify({"status": "error", "message": "loc parameter required"}), 400
    try:
        from geopy.geocoders import Nominatim

        geolocator = Nominatim(user_agent="telco_mapper_app")
        location = geolocator.geocode(loc)
        if location:
            return jsonify(
                {
                    "status": "success",
                    "lat": location.latitude,
                    "lon": location.longitude,
                }
            )
        else:
            return jsonify({"status": "not_found"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})


@app.route("/api/georeference/poles")
def api_georeference_poles():
    """Return poles from POLE_STATE formatted for the georeference workflow."""
    tags = POLE_STATE.get("tags", [])
    out = []
    for t in tags:
        out.append(
            {
                "id": t.get("pole_id", 0),
                "name": t.get("name", ""),
                "cx": t.get("cx", 0),
                "cy": t.get("cy", 0),
            }
        )
    return jsonify({"status": "success", "poles": out})


@app.route("/api/georeference/process", methods=["POST"])
def api_georeference_process():
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"status": "error", "message": "Request body required"}), 400

    anchors = data.get("anchors", [])
    poles_data = data.get("poles", [])

    if len(anchors) != 4:
        return jsonify(
            {"status": "error", "message": "Exactly 4 anchor points required"}
        ), 400

    if not poles_data:
        return jsonify({"status": "error", "message": "No poles provided"}), 400

    # Bounds from GPS anchors
    lats = [a[0] for a in anchors]
    lons = [a[1] for a in anchors]
    map_p1 = (min(lats), min(lons))
    map_p2 = (max(lats), max(lons))

    # Bounds from CAD poles
    cxs = [p.get("cx", 0) for p in poles_data]
    cys = [p.get("cy", 0) for p in poles_data]
    cad_p1 = (min(cxs), min(cys))
    cad_p2 = (max(cxs), max(cys))

    mapped = []
    for p in poles_data:
        lat, lon = transform_coordinate(
            (p["cx"], p["cy"]), cad_p1, cad_p2, map_p1, map_p2
        )
        mapped.append(
            {
                "id": p.get("id", 0),
                "name": p.get("name", ""),
                "lat": round(lat, 6),
                "lon": round(lon, 6),
                "cad_x": p["cx"],
                "cad_y": p["cy"],
            }
        )

    return jsonify({"status": "success", "poles": mapped})


# ─────────────────────────────────────────────────────────────────────────────
# PUBLIC INTEGRATION API  —  /api/v1/
# ─────────────────────────────────────────────────────────────────────────────


def _v1_ok(data: Any, status: int = 200):
    return jsonify({"ok": True, "data": data}), status


def _v1_err(message: str, status: int = 400):
    return jsonify({"ok": False, "error": message}), status


@public_api.route("/health", methods=["GET"])
def v1_health():
    try:
        import easyocr  # noqa

        engine_ok = True
    except ImportError:
        engine_ok = False
    return _v1_ok(
        {
            "service": "strand-identifier",
            "version": "1.0.0",
            "engine": "easyocr",
            "engine_ready": engine_ok,
        }
    )


@public_api.route("/status", methods=["GET"])
def v1_status():
    return _v1_ok(
        {
            "dxf_path": state.get("dxf_path"),
            "ocr": {
                "status": state.get("status", "idle"),
                "progress": state.get("progress", 0),
                "total": state.get("total", 0),
                "step": state.get("step", 0),
                "step_label": state.get("step_label", ""),
                "error": state.get("error"),
            },
            "equipment": {
                "status": SCAN_STATE.get("status", "idle"),
                "progress": SCAN_STATE.get("progress", 0),
                "total": SCAN_STATE.get("total", 0),
                "count": len(SCAN_STATE.get("shapes", [])),
                "error": SCAN_STATE.get("error"),
            },
            "poles": {
                "status": POLE_STATE.get("status", "idle"),
                "layer": POLE_STATE.get("layer"),
                "progress": POLE_STATE.get("progress", 0),
                "total": POLE_STATE.get("total", 0),
                "count": len(POLE_STATE.get("tags", [])),
                "error": POLE_STATE.get("error"),
            },
        }
    )


@public_api.route("/ocr/results", methods=["GET"])
def v1_ocr_results():
    if state.get("status") == "idle":
        return _v1_err("No OCR run has been started yet.", 404)
    if state.get("status") == "processing":
        return _v1_err(
            "OCR is still processing. Poll /api/v1/status for progress.", 202
        )
    if state.get("status") == "error":
        return _v1_err(f"OCR pipeline failed: {state.get('error')}", 500)

    include_crops = request.args.get("include_crops", "false").lower() == "true"
    filter_review = request.args.get("needs_review", "").lower()
    raw_results = list(state.get("results", []))

    if filter_review == "true":
        raw_results = [r for r in raw_results if r.get("needs_review")]
    elif filter_review == "false":
        raw_results = [r for r in raw_results if not r.get("needs_review")]

    output = []
    total_sum = 0
    for r in raw_results:
        final = r.get("corrected_value") or r.get("value", "")
        try:
            total_sum += int(final)
        except (ValueError, TypeError):
            pass
        output.append(
            {
                "digit_id": r.get("digit_id"),
                "value": r.get("value"),
                "corrected_value": r.get("corrected_value"),
                "final_value": final,
                "confidence": r.get("confidence"),
                "needs_review": r.get("needs_review"),
                "center_x": r.get("center_x"),
                "center_y": r.get("center_y"),
                "bbox": r.get("bbox"),
                "manual": r.get("manual", False),
                "crop_b64": r.get("crop_b64") if include_crops else None,
            }
        )

    return _v1_ok(
        {
            "dxf_path": state.get("dxf_path"),
            "count": len(output),
            "sum": total_sum,
            "results": output,
        }
    )


@public_api.route("/ocr/segments", methods=["GET"])
def v1_ocr_segments():
    raw_segs = state.get("segments", [])
    segments = [
        {
            "x1": s.x1,
            "y1": s.y1,
            "x2": s.x2,
            "y2": s.y2,
            "is_hatch": getattr(s, "is_hatch", False),
        }
        for s in raw_segs
    ]
    return _v1_ok(
        {
            "dxf_path": state.get("dxf_path"),
            "count": len(segments),
            "segments": segments,
        }
    )


@public_api.route("/poles", methods=["GET"])
def v1_poles():
    status = POLE_STATE.get("status", "idle")
    if status == "idle":
        return _v1_err("No pole scan has been started yet.", 404)
    if status == "processing":
        return _v1_err("Pole scan is still running. Poll /api/v1/status.", 202)
    if status == "error":
        return _v1_err(f"Pole scan failed: {POLE_STATE.get('error')}", 500)

    include_crops = request.args.get("include_crops", "false").lower() == "true"
    filter_review = request.args.get("needs_review", "").lower()
    filter_source = request.args.get("source", "").lower()

    tags = list(POLE_STATE.get("tags", []))

    if filter_review == "true":
        tags = [t for t in tags if t.get("needs_review")]
    elif filter_review == "false":
        tags = [t for t in tags if not t.get("needs_review")]
    if filter_source in ("text", "mtext", "stroke"):
        tags = [t for t in tags if t.get("source") == filter_source]

    output = [
        {
            "pole_id": t.get("pole_id"),
            "name": t.get("name"),
            "cx": t.get("cx"),
            "cy": t.get("cy"),
            "bbox": t.get("bbox"),
            "layer": t.get("layer"),
            "source": t.get("source"),
            "ocr_conf": t.get("ocr_conf"),
            "needs_review": t.get("needs_review"),
            "crop_b64": t.get("crop_b64") if include_crops else None,
            "map_latitude": t.get("map_latitude"),
            "map_longitude": t.get("map_longitude"),
        }
        for t in tags
    ]

    return _v1_ok(
        {
            "dxf_path": state.get("dxf_path"),
            "layer": POLE_STATE.get("layer"),
            "count": len(output),
            "poles": output,
        }
    )


@public_api.route("/poles/georeference", methods=["PATCH"])
def v1_poles_georeference():
    body = request.get_json(silent=True)
    if not body or "poles" not in body:
        return _v1_err("Request body must contain a 'poles' array.", 400)

    tags = POLE_STATE.get("tags", [])
    tag_map = {
        key: t
        for t in tags
        if (key := _pole_id_key(t.get("pole_id"))) is not None
    }
    updated = 0
    added = 0  # Track new NPTs

    for p in body["poles"]:
        pid = p.get("pole_id")
        pid_key = _pole_id_key(pid)
        lat = p.get("map_latitude")
        lon = p.get("map_longitude")
        name = p.get("name", "NPT")
        layer = p.get("layer", "geotool_npt")
        source = p.get("source", "geotool_npt")

        if lat is None or lon is None:
            continue

        if pid_key is not None and pid_key in tag_map:
            # Update existing pole
            tag_map[pid_key]["map_latitude"] = lat
            tag_map[pid_key]["map_longitude"] = lon
            tag_map[pid_key]["cx"] = p.get("cad_x", tag_map[pid_key].get("cx", 0))
            tag_map[pid_key]["cy"] = p.get("cad_y", tag_map[pid_key].get("cy", 0))
            tag_map[pid_key]["layer"] = layer or tag_map[pid_key].get("layer", "geotool_npt")
            tag_map[pid_key]["source"] = source or tag_map[pid_key].get("source", "geotool_npt")
            updated += 1
        else:
            # UPSERT: Insert newly discovered NPTs from the geotool
            new_pole = {
                "pole_id": pid,
                "name": name,
                "cx": p.get("cad_x", 0),
                "cy": p.get("cad_y", 0),
                "bbox": [0, 0, 0, 0],
                "layer": layer or "geotool_npt",
                "source": source or "geotool_npt",
                "ocr_conf": 1.0,
                "needs_review": False,
                "crop_b64": None,
                "map_latitude": lat,
                "map_longitude": lon,
            }
            tags.append(new_pole)
            if pid_key is not None:
                tag_map[pid_key] = new_pole
            added += 1

    POLE_STATE["tags"] = tags
    return _v1_ok({"updated": updated, "added": added})


@public_api.route("/equipment", methods=["GET"])
def v1_equipment():
    status = SCAN_STATE.get("status", "idle")
    if status == "idle":
        return _v1_err("No equipment scan has been started yet.", 404)
    if status == "processing":
        return _v1_err("Equipment scan is still running. Poll /api/v1/status.", 202)
    if status == "error":
        return _v1_err(f"Equipment scan failed: {SCAN_STATE.get('error')}", 500)

    filter_kind = request.args.get("kind", "").lower()
    filter_layer = request.args.get("layer", "")

    all_shapes = SCAN_STATE.get("shapes", [])
    summary: Dict[str, int] = {}
    for s in all_shapes:
        summary[s["kind"]] = summary.get(s["kind"], 0) + 1

    filtered = list(all_shapes)
    if filter_kind:
        filtered = [s for s in filtered if s.get("kind") == filter_kind]
    if filter_layer:
        filtered = [s for s in filtered if s.get("layer") == filter_layer]

    return _v1_ok(
        {
            "dxf_path": state.get("dxf_path"),
            "count": len(filtered),
            "summary": summary,
            "shapes": [
                {
                    "shape_id": s.get("shape_id"),
                    "kind": s.get("kind"),
                    "cx": s.get("cx"),
                    "cy": s.get("cy"),
                    "bbox": s.get("bbox"),
                    "layer": s.get("layer"),
                }
                for s in filtered
            ],
        }
    )


@public_api.route("/cable_spans", methods=["GET"])
def v1_cable_spans():
    dxf_path = state.get("dxf_path")
    if not dxf_path:
        return _v1_err("No DXF has been loaded.", 404)

    include_segs = request.args.get("include_segments", "false").lower() == "true"

    try:
        body = _span_response(dxf_path)
        body["dxf_path"] = dxf_path
        if not include_segs:
            for s in body.get("spans", []):
                s["segments"] = None
        return _v1_ok(body)

    except Exception as exc:
        import traceback

        traceback.print_exc()
        return _v1_err(str(exc), 500)


@public_api.route("/cable_spans/derive", methods=["POST"])
def v1_cable_spans_derive():
    """Re-derive spans after the pole set or an OCR value changed.

    Pole edits and digit corrections both move span boundaries, so the client
    posts its current pole list and corrections here rather than patching spans
    locally. Everything downstream — canvas, PDF, both exports — then reads one
    derivation instead of drifting apart.
    """
    body = request.get_json(silent=True) or {}
    # Restored sessions arrive after a server restart, when no DXF is loaded
    # yet — the client names the drawing it wants derived.
    dxf_path = body.get("dxf_path") or state.get("dxf_path")
    if not dxf_path:
        return _v1_err("No DXF has been loaded.", 404)
    if dxf_path != state.get("dxf_path"):
        state["dxf_path"] = dxf_path

    poles = body.get("poles")
    corrections = body.get("corrections") or {}

    try:
        if isinstance(poles, list):
            POLE_STATE["tags"] = poles
        for r in state.get("results", []):
            did = str(r.get("digit_id"))
            if did in corrections and corrections[did] is not None:
                r["corrected_value"] = corrections[did]

        result, cable_layers = derive_node_spans(dxf_path)
        if result is None:
            return _v1_ok(
                {
                    "dxf_path": dxf_path,
                    "cable_layers": [],
                    "count": 0,
                    "spans": [],
                    "status": "no_cable_layer",
                }
            )

        payload = result.to_dict()
        payload["dxf_path"] = dxf_path
        payload["cable_layers"] = cable_layers
        payload["status"] = "ok" if result.ok else "blocked"
        return _v1_ok(payload)

    except Exception as exc:
        import traceback

        traceback.print_exc()
        return _v1_err(str(exc), 500)


@public_api.route("/export/ocr", methods=["POST"])
def v1_export_ocr():
    if state.get("status") != "done":
        return _v1_err("OCR must complete before exporting.", 400)
    body = request.get_json(silent=True) or {}
    corrections = body.get("corrections", {})
    for r in state.get("results", []):
        did = str(r["digit_id"])
        if did in corrections and corrections[did] is not None:
            r["corrected_value"] = corrections[did]
    path, err = export_excel(state["results"], state["dxf_path"])
    if err:
        return _v1_err(err, 500)
    return _v1_ok({"download_url": f"/api/download?file={path}", "path": path})


@public_api.route("/export/poles", methods=["POST"])
def v1_export_poles():
    body = request.get_json(silent=True) or {}
    tags = body.get("poles", POLE_STATE.get("tags", []))
    if not tags:
        return _v1_err("No pole tags to export. Run a scan first.", 400)
    overrides = body.get("overrides", {})
    export_tags = []
    for t in tags:
        pid_str = str(t.get("pole_id", ""))
        entry = dict(t)
        if pid_str in overrides:
            entry["name"] = overrides[pid_str]
        export_tags.append(entry)
    dxf_path = state.get("dxf_path") or "pole_export"
    path, err = export_poles_excel(export_tags, dxf_path)
    if err:
        return _v1_err(err, 500)
    return _v1_ok({"download_url": f"/api/download?file={path}", "path": path})


# ─────────────────────────────────────────────────────────────────────────────
# ASBUILT IQ EXPORT  —  /api/v1/asbuilt/
# ─────────────────────────────────────────────────────────────────────────────


@public_api.route("/asbuilt/sites", methods=["GET"])
def v1_asbuilt_sites():
    """Proxy to AsBuilt IQ API — list all sites (areas)."""
    try:
        from app_python.services.asbuilt_api import get_sites

        sites = get_sites()
        return _v1_ok(sites)
    except Exception as e:
        return _v1_err(str(e), 502)


@public_api.route("/asbuilt/sites/<int:area_id>/nodes", methods=["GET"])
def v1_asbuilt_nodes(area_id: int):
    """Proxy to AsBuilt IQ API — list nodes under an area."""
    try:
        from app_python.services.asbuilt_api import get_nodes

        nodes = get_nodes(area_id)
        return _v1_ok(nodes)
    except Exception as e:
        return _v1_err(str(e), 502)


@public_api.route("/asbuilt/node/<int:node_id>", methods=["GET"])
def v1_asbuilt_node(node_id: int):
    """Proxy to AsBuilt IQ API — verify node state."""
    try:
        from app_python.services.asbuilt_api import get_node

        node = get_node(node_id)
        return _v1_ok(node)
    except Exception as e:
        return _v1_err(str(e), 502)


@public_api.route("/asbuilt/subcontractors", methods=["GET"])
def v1_asbuilt_subcontractors():
    """Proxy to AsBuilt IQ API — list all subcontractors."""
    try:
        from app_python.services.asbuilt_api import get_subcontractors

        subcontractors = get_subcontractors()
        return _v1_ok(subcontractors)
    except Exception as e:
        return _v1_err(str(e), 502)


@public_api.route("/asbuilt/teams", methods=["GET"])
def v1_asbuilt_teams():
    """Proxy to AsBuilt IQ API — list teams, optionally filtered by subcontractor."""
    try:
        from app_python.services.asbuilt_api import get_teams

        subcontractor_id = request.args.get("subcontractor_id", type=int)
        teams = get_teams(subcontractor_id)
        return _v1_ok(teams)
    except Exception as e:
        return _v1_err(str(e), 502)


@public_api.route("/asbuilt/import", methods=["POST"])
def v1_asbuilt_import():
    """
    Import poles + spans to the AsBuilt IQ API.

    Request Body
    ------------
    {
        "node_id":   "TY1401",            # Required — VARCHAR node identifier
        "node_name": "MONTEVISTA SUBD.",  # Required — node name
        "area_id":   1,                   # Required — area/site database ID
        "region":    "CALABARZON",        # Optional
        "province":  "LAGUNA",            # Optional
        "city":      "STA. ROSA",         # Optional
        "barangay_code": "043428001",     # Optional
        "barangay_name": "Balibago",       # Optional
        "poles": [                         # Required
            {
                "pole_code": "PL-001",
                "latitude":  14.539770,
                "longitude": 121.109219,
                "region":    "CALABARZON",
                "province":  "LAGUNA",
                "city":      "STA. ROSA",
                "barangay_code": "043428001",
                "barangay_name": "Balibago"
            }
        ],
        "spans": [                         # Optional
            {
                "from_pole_code": "PL-001",
                "to_pole_code":   "PL-002",
                "strand_length":  50.5,
                "number_of_runs": 1,
                "components": {
                    "node": 2, "amplifier": 1, "extender": 0,
                    "tsc": 1, "powersupply": 0, "ps_housing": 0
                }
            }
        ]
    }

    Response
    --------
    {
        "ok": true,
        "data": {
            "message": "AsBuilt import completed.",
            "data": {
                "node": { ... },
                "poles_created": [...],
                "poles_updated": [...],
                "spans_created": [...],
                "spans_updated": [...],
                "total_poles": 3,
                "total_spans": 2,
                "errors": []
            }
        }
    }
    """
    body = request.get_json(silent=True) or {}
    node_id = body.get("node_id")
    if not node_id:
        return _v1_err("node_id (VARCHAR) is required", 400)
    node_name = body.get("node_name")
    if not node_name:
        return _v1_err("node_name is required", 400)

    poles = body.get("poles")
    if not poles or not isinstance(poles, list) or len(poles) == 0:
        return _v1_err("poles array is required", 400)

    def _safe_float(value: Any) -> Optional[float]:
        try:
            num = float(value)
        except (TypeError, ValueError):
            return None
        if not math.isfinite(num):
            return None
        return num

    # Normalize and validate poles before sending them to the remote API.
    cleaned_poles = []
    seen_pole_codes = set()
    skipped_poles = 0
    for raw_pole in poles:
        if not isinstance(raw_pole, dict):
            skipped_poles += 1
            continue

        pole_code = (raw_pole.get("pole_code") or "").strip().upper()
        latitude = _safe_float(raw_pole.get("latitude"))
        longitude = _safe_float(raw_pole.get("longitude"))
        if not pole_code or latitude is None or longitude is None:
            skipped_poles += 1
            continue
        if pole_code in seen_pole_codes:
            skipped_poles += 1
            continue

        seen_pole_codes.add(pole_code)
        cleaned_pole = {
            "pole_code": pole_code,
            "latitude": latitude,
            "longitude": longitude,
        }

        pole_index = raw_pole.get("pole_index")
        if pole_index is not None:
            cleaned_pole["pole_index"] = pole_index

        for field in ("region", "province", "city", "barangay_code", "barangay_name"):
            val = raw_pole.get(field)
            if val not in (None, ""):
                cleaned_pole[field] = val

        cleaned_poles.append(cleaned_pole)

    if not cleaned_poles:
        return _v1_err(
            "No valid poles were available for AsBuilt export. "
            "Each pole needs a name/code plus valid latitude and longitude.",
            400,
        )

    # Build payload — frontend already handles format, area data, and dedup
    payload: dict[str, Any] = {
        "node_id": node_id,
        "node_name": node_name,
        "area_id": body.get("area_id"),
        "poles": cleaned_poles,
    }

    # Pass through optional area fields
    for field in (
        "region",
        "province",
        "city",
        "barangay_code",
        "barangay_name",
        "subcontractor_id",
        "team_id",
    ):
        val = body.get(field)
        if val:
            payload[field] = val

    # Attach spans with component defaults
    spans = body.get("spans")
    if spans and isinstance(spans, list):
        span_defaults = {
            "node": 0,
            "amplifier": 0,
            "extender": 0,
            "tsc": 0,
            "powersupply": 0,
            "ps_housing": 0,
        }
        cleaned = []
        skipped_spans = 0
        for s in spans:
            if not isinstance(s, dict):
                skipped_spans += 1
                continue
            frm = (s.get("from_pole_code") or "").strip().upper()
            to = (s.get("to_pole_code") or "").strip().upper()
            if (
                not frm
                or not to
                or frm == to
                or frm not in seen_pole_codes
                or to not in seen_pole_codes
            ):
                skipped_spans += 1
                continue
            comp = dict(span_defaults)
            if isinstance(s.get("components"), dict):
                comp.update(s["components"])
            strand_length = _safe_float(s.get("strand_length"))
            try:
                from_index = int(s.get("from_pole_index"))
                to_index = int(s.get("to_pole_index"))
                number_of_runs = int(s.get("number_of_runs", 1))
            except (TypeError, ValueError):
                skipped_spans += 1
                continue
            if from_index < 1 or to_index < 1 or number_of_runs < 1:
                skipped_spans += 1
                continue
            cleaned.append(
                {
                    "from_pole_code": frm,
                    "to_pole_code": to,
                    "from_pole_index": from_index,
                    "to_pole_index": to_index,
                    "strand_length": strand_length if strand_length is not None else 0.0,
                    "number_of_runs": number_of_runs,
                    "components": comp,
                }
            )
        if cleaned:
            payload["spans"] = cleaned

    payload_summary = {
        "node_id": node_id,
        "area_id": body.get("area_id"),
        "poles": len(cleaned_poles),
        "spans": len(payload.get("spans", [])),
        "skipped_poles": skipped_poles,
        "skipped_spans": skipped_spans if spans and isinstance(spans, list) else 0,
    }

    try:
        from app_python.services.asbuilt_api import import_data

        result = import_data(payload)
        return _v1_ok(result)
    except Exception as e:
        status = 502
        detail = str(e)
        resp = getattr(e, "response", None)
        if resp is not None:
            status = resp.status_code
            try:
                detail = resp.json()
            except Exception:
                detail = str(e)
        print(f"[asbuilt] Import failed. Summary={payload_summary} Error={detail}")
        return _v1_err(
            f"AsBuilt API error ({status}): {detail}. Payload summary: {payload_summary}",
            status,
        )


@public_api.route("/asbuilt/import-by-sequence", methods=["POST"])
def v1_asbuilt_import_by_sequence():
    body = request.get_json(silent=True) or {}
    node_id = body.get("node_id")
    if not node_id:
        return _v1_err("node_id (VARCHAR) is required", 400)
    node_name = body.get("node_name")
    if not node_name:
        return _v1_err("node_name is required", 400)

    poles = body.get("poles")
    if not poles or not isinstance(poles, list) or len(poles) == 0:
        return _v1_err("poles array is required", 400)

    def _safe_float(value: Any) -> Optional[float]:
        try:
            num = float(value)
        except (TypeError, ValueError):
            return None
        if not math.isfinite(num):
            return None
        return num

    cleaned_poles = []
    seen_pole_indexes = set()
    skipped_poles = 0
    for raw_pole in poles:
        if not isinstance(raw_pole, dict):
            skipped_poles += 1
            continue

        pole_index = str(raw_pole.get("pole_index") or "").strip().upper()
        pole_code = str(raw_pole.get("pole_code") or "").strip().upper()
        latitude = _safe_float(raw_pole.get("lat", raw_pole.get("latitude")))
        longitude = _safe_float(raw_pole.get("lng", raw_pole.get("longitude")))
        if not pole_index or not pole_code or latitude is None or longitude is None:
            skipped_poles += 1
            continue
        if pole_index in seen_pole_indexes:
            skipped_poles += 1
            continue

        seen_pole_indexes.add(pole_index)
        cleaned_pole = {
            "pole_index": pole_index,
            "pole_code": pole_code,
            "lat": latitude,
            "lng": longitude,
        }

        for field in ("region", "province", "city"):
            val = raw_pole.get(field)
            if val not in (None, ""):
                cleaned_pole[field] = val

        cleaned_poles.append(cleaned_pole)

    if not cleaned_poles:
        return _v1_err(
            "No valid poles were available for AsBuilt export. "
            "Each pole needs a unique pole_index, pole_code, latitude, and longitude.",
            400,
        )

    payload: dict[str, Any] = {
        "node_id": node_id,
        "node_name": node_name,
        "area_id": body.get("area_id"),
        "poles": cleaned_poles,
    }

    for field in (
        "region",
        "province",
        "city",
        "barangay_name",
        "subcontractor_id",
        "team_id",
    ):
        val = body.get(field)
        if val:
            payload[field] = val

    spans = body.get("spans")
    skipped_spans = 0
    span_defaults = {
        "node": 0,
        "amplifier": 0,
        "extender": 0,
        "tsc": 0,
        "powersupply": 0,
        "ps_housing": 0,
    }
    component_totals = dict(span_defaults)
    if spans and isinstance(spans, list):
        cleaned = []
        for s in spans:
            if not isinstance(s, dict):
                skipped_spans += 1
                continue
            frm = str(s.get("from_pole_index") or "").strip().upper()
            to = str(s.get("to_pole_index") or "").strip().upper()
            if (
                not frm
                or not to
                or frm == to
                or frm not in seen_pole_indexes
                or to not in seen_pole_indexes
            ):
                skipped_spans += 1
                continue
            comp = dict(span_defaults)
            if isinstance(s.get("components"), dict):
                comp.update(s["components"])
            for key in span_defaults:
                try:
                    count = int(comp.get(key) or 0)
                except (TypeError, ValueError):
                    count = 0
                if count < 0:
                    count = 0
                comp[key] = count
            strand_length = _safe_float(s.get("strand_length"))
            try:
                number_of_runs = int(s.get("number_of_runs", 1))
            except (TypeError, ValueError):
                skipped_spans += 1
                continue
            if strand_length is None or number_of_runs < 1:
                skipped_spans += 1
                continue
            for key in span_defaults:
                component_totals[key] += comp[key]
            cleaned.append(
                {
                    "from_pole_index": frm,
                    "to_pole_index": to,
                    "strand_length": strand_length,
                    "number_of_runs": number_of_runs,
                    "components": comp,
                }
            )
        if cleaned:
            payload["spans"] = cleaned

    payload_summary = {
        "node_id": node_id,
        "area_id": body.get("area_id"),
        "poles": len(cleaned_poles),
        "spans": len(payload.get("spans", [])),
        "skipped_poles": skipped_poles,
        "skipped_spans": skipped_spans,
        "components": component_totals,
    }

    try:
        from app_python.services.asbuilt_api import get_nodes, import_data_by_sequence

        area_id = body.get("area_id")
        if area_id:
            nodes_response = get_nodes(int(area_id))
            nodes_payload = (
                nodes_response.get("nodes")
                if isinstance(nodes_response, dict)
                else nodes_response
            )
            if isinstance(nodes_payload, dict):
                nodes_payload = nodes_payload.get("nodes", [])
            if not isinstance(nodes_payload, list):
                nodes_payload = []

            normalized_node_id = str(node_id).strip().upper()
            duplicate_node = next(
                (
                    node
                    for node in nodes_payload
                    if str(node.get("node_id", "")).strip().upper()
                    == normalized_node_id
                ),
                None,
            )
            if duplicate_node:
                existing_label = (
                    duplicate_node.get("full_label")
                    or duplicate_node.get("name")
                    or duplicate_node.get("node_id")
                )
                existing_db_id = duplicate_node.get("id")
                return _v1_err(
                    "Node ID already exists in the selected site before import. "
                    f"Existing node: {existing_label}"
                    f"{f' (database id {existing_db_id})' if existing_db_id else ''}. "
                    "Delete that partial/old node first or use a fresh exact Node ID.",
                    409,
                )

        print(f"[asbuilt-sequence] Sending import. Summary={payload_summary}")
        result = import_data_by_sequence(payload)
        return _v1_ok(result)
    except Exception as e:
        status = 502
        detail = str(e)
        resp = getattr(e, "response", None)
        if resp is not None:
            status = resp.status_code
            try:
                detail = resp.json()
            except Exception:
                detail = str(e)
        print(f"[asbuilt-sequence] Import failed. Summary={payload_summary} Error={detail}")
        return _v1_err(
            f"AsBuilt sequence API error ({status}): {detail}. Payload summary: {payload_summary}",
            status,
        )


# ─────────────────────────────────────────────────────────────────────────────
# STARTUP
# ─────────────────────────────────────────────────────────────────────────────


def _prewarm_ocr():
    try:
        print("[startup] Pre-warming strand recognizer (EasyOCR)...")
        from app_python.services.strand_recognizer import prewarm as _sr_prewarm
        _sr_prewarm()
        # Keep the legacy singleton warm too (used by _easyocr_on_prepared)
        _load_easyocr()
        print("[startup] Strand recognizer ready.")
    except Exception as e:
        print(f"[startup] Strand recognizer pre-warm failed: {e}")

    if os.environ.get("REMOTE_TROCR_URL"):
        print("[startup] Pole TrOCR pre-warm skipped (remote mode)")
    else:
        try:
            from app_python.services.pole_ocr import _load_model as _load_pole

            print("[startup] Pre-warming pole TrOCR model...")
            _load_pole()
            print("[startup] Pole TrOCR ready.")
        except Exception as e:
            print(f"[startup] Pole TrOCR pre-warm failed: {e}")


app.register_blueprint(public_api)


def _start_background_prewarm():
    if os.environ.get("DISABLE_OCR_PREWARM", "").lower() in ("1", "true", "yes"):
        print("[startup] OCR pre-warm disabled.")
        return

    threading.Thread(target=_prewarm_ocr, name="ocr-prewarm", daemon=True).start()


_start_background_prewarm()

_geotool_proc = None


def _start_geotool():
    global _geotool_proc
    project_root = Path(__file__).parent
    _geotool_proc = subprocess.Popen(
        [sys.executable, "-m", "geotool", "--port", "8000"],
        cwd=str(project_root),
    )
    print(f"  GeoTool server -> http://localhost:8000")
    atexit.register(lambda: _geotool_proc and _geotool_proc.terminate())


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--port", type=int, default=5000)
    parser.add_argument("--dev", action="store_true")
    args = parser.parse_args()

    print(f"\n{'=' * 50}")
    print(f"  CAD OCR – Flask Backend  (EasyOCR engine)")
    print(f"  http://localhost:{args.port}")
    if args.dev:
        print(f"  React dev server: http://localhost:5173")
    print(f"{'=' * 50}\n")

    _start_geotool()
    app.run(host="localhost", port=args.port, debug=True, threaded=True)


def _can_bind_port(host: str, port: int) -> bool:
    try:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
            sock.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
            sock.bind((host, port))
        return True
    except OSError:
        return False


def _resolve_runtime_port() -> int:
    raw_port = os.environ.get("PORT")
    if raw_port:
        try:
            return int(raw_port)
        except ValueError:
            print(f"[startup] Invalid PORT={raw_port!r}; falling back to default.")

    default_port = 5000
    fallback_port = 5050

    if _can_bind_port("0.0.0.0", default_port):
        return default_port

    print(
        f"[startup] Port {default_port} is unavailable on this machine. "
        f"Falling back to {fallback_port}."
    )
    return fallback_port


if __name__ == "__main__":
    port = _resolve_runtime_port()
    print(f"\n{'=' * 50}")
    print(f"  CAD OCR – Flask Backend  (EasyOCR engine)")
    print(f"  http://localhost:{port}")
    print(f"{'=' * 50}\n")
    _start_geotool()
    app.run(host="0.0.0.0", port=port)
